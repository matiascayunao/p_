import json
import re
import unicodedata
from datetime import date, datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from .excel_utils import build_excel_sectores,  build_excel_plantilla_carga_masiva
from django.db.models import Sum, Q, Case, When, IntegerField, F, Value
from django.views.decorators.http import require_GET, require_POST, require_http_methods
from django.contrib import messages
from django.apps import apps



from .forms import (
    CrearSector, CrearUbicacion, CrearPiso, CrearLugar,
    CrearObjetoLugar, CrearTipoLugar, CrearCategoriaObjeto,
    CrearObjeto, CrearTipoObjeto, CrearHistorico,
    EditarSector, EditarUbicacion, EditarPiso, EditarLugar,
    EditarTipoLugar, EditarCategoria, EditarObjeto,
    EditarTipoObjeto, EditarObjetoLugar, EditarHistorico,
    EstructuraCompletaForm, ObjetoLugarFilaFormSet, UploadExcelForm
)

from .models import (
    Sector, Ubicacion, Piso, Lugar, TipoLugar,
    CategoriaObjeto, Objeto, TipoObjeto,
    ObjetoLugar, HistoricoObjeto, TipoLugarObjetoTipico,
    AreaMapa
)

import openpyxl 
from openpyxl import load_workbook



# -------------------
# AUTH
# -------------------   

@login_required
def descargar_excel_sectores(request):
    ubicaciones = Ubicacion.objects.select_related("sector").order_by("sector__sector","ubicacion")
    xlsx_bytes = build_excel_sectores(ubicaciones)

    hoy =date.today().strftime("%d-%m-%Y")
    filename= f'PUERTO_{hoy}.xlsx'
    response = HttpResponse(xlsx_bytes, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"]= f'attachment; filename= "{filename}"'
    return response

@login_required
@transaction.atomic
def crear_estructura(request):
    if request.method == "POST":
        form = EstructuraCompletaForm(request.POST)
        objetos_formset = ObjetoLugarFilaFormSet(request.POST, prefix="obj")

        if form.is_valid() and objetos_formset.is_valid():
            cd = form.cleaned_data

            # -----------------------
            # 1) SECTOR
            # -----------------------
            sector = cd.get("sector_existente")
            sector_nuevo = (cd.get("sector_nuevo") or "").strip()
            if not sector and sector_nuevo:
                sector, _ = Sector.objects.get_or_create(sector=sector_nuevo)

            # -----------------------
            # 2) UBICACIÓN
            # -----------------------
            ubicacion = cd.get("ubicacion_existente")
            ubicacion_nueva = (cd.get("ubicacion_nueva") or "").strip()
            if not ubicacion and ubicacion_nueva:
                ubicacion, _ = Ubicacion.objects.get_or_create(
                    ubicacion=ubicacion_nueva,
                    sector=sector,
                )

            # -----------------------
            # 3) PISO
            # -----------------------
            piso = cd.get("piso_existente")
            piso_nuevo = cd.get("piso_nuevo")
            if not piso and piso_nuevo not in (None, ""):
                piso, _ = Piso.objects.get_or_create(
                    piso=piso_nuevo,
                    ubicacion=ubicacion,
                )

            # -----------------------
            # 4) TIPO DE LUGAR
            # -----------------------
            tipo_lugar = cd.get("tipo_lugar_existente")
            tipo_lugar_nuevo = (cd.get("tipo_lugar_nuevo") or "").strip()
            if not tipo_lugar and tipo_lugar_nuevo:
                tipo_lugar, _ = TipoLugar.objects.get_or_create(
                    tipo_de_lugar=tipo_lugar_nuevo
                )

            # -----------------------
            # 5) LUGAR
            # -----------------------
            lugar = cd.get("lugar_existente")
            lugar_nuevo = (cd.get("lugar_nuevo")or "").strip()

            if not lugar:
                lugar = Lugar.objects.create(nombre_del_lugar = lugar_nuevo, piso=piso, lugar_tipo_lugar = tipo_lugar,)

            # -----------------------
            # 6) OBJETOS DEL LUGAR
            # -----------------------
            for f in objetos_formset:
                if not f.cleaned_data:
                    continue
                if f.cleaned_data.get("__empty__"):
                    continue

                row = f.cleaned_data

                # --- Categoría ---
                categoria = row.get("categoria_existente")
                if not categoria:
                    nombre_cat = (row.get("categoria_nueva") or "").strip()
                    categoria, _ = CategoriaObjeto.objects.get_or_create(
                        nombre_de_categoria=nombre_cat
                    )

                # --- Objeto ---
                objeto = row.get("objeto_existente")
                if not objeto:
                    nombre_obj = (row.get("objeto_nuevo") or "").strip()
                    objeto, _ = Objeto.objects.get_or_create(
                        nombre_del_objeto=nombre_obj,
                        objeto_categoria=categoria,
                    )

                # --- Tipo de objeto ---
                tipo_obj = row.get("tipo_objeto_existente")
                if not tipo_obj:
                    marca = (row.get("marca") or "").strip()
                    material = (row.get("material") or "").strip()

                    if not marca:
                        marca = "Sin marca"
                    if not material:
                        material=  "Sin material"

                    tipo_obj, _ = TipoObjeto.objects.get_or_create(
                        objeto=objeto,
                        marca=marca,
                        material=material,
                    )

                cantidad = row.get("cantidad") or 0
                estado = row.get("estado") or "B"
                detalle = (row.get("detalle") or "").strip()

                ObjetoLugar.objects.create(
                    lugar=lugar,
                    tipo_de_objeto=tipo_obj,
                    cantidad=cantidad,
                    estado=estado,
                    detalle=detalle,
                )

            # Redirige al detalle del lugar recién creado
            return redirect("detalle_lugar", lugar_id=lugar.id)

    else:
        form = EstructuraCompletaForm()
        objetos_formset = ObjetoLugarFilaFormSet(prefix="obj")

    return render(
        request,
        "crear_estructura.html",  # cambia este nombre si tu template es otro
        {
            "form": form,
            "objetos_formset": objetos_formset,
        },
    )


def signup(request):
    if request.method == "GET":
        return render(request, "signup.html", {"form": UserCreationForm})
    if request.POST["password1"] != request.POST["password2"]:
        return render(
            request,
            "signup.html",
            {"form": UserCreationForm, "error": "Las contraseñas no coinciden"},
        )

    try:
        user = User.objects.create_user(
            username=request.POST["username"],
            password=request.POST["password1"],
        )
        user.save()
        login(request, user)
        return redirect("home")
    except Exception:
        return render(
            request,
            "signup.html",
            {"form": UserCreationForm, "error": "Usuario ya existe"},
        )


def signin(request):
    if request.method == "GET":
        return render(request, "signin.html", {"form": AuthenticationForm})

    user = authenticate(
        request,
        username=request.POST["username"],
        password=request.POST["password"],
    )
    if user is None:
        return render(
            request,
            "signin.html",
            {
                "form": AuthenticationForm,
                "error": "Nombre o contraseña incorrectos",
            },
        )

    login(request, user)
    return redirect("home")


@login_required
def signout(request):
    logout(request)
    return redirect("signin")


@login_required
def home(request):
    vista_actual = request.GET.get("vista", "sectores")
    sectores = Sector.objects.order_by("sector")
    ubicaciones = Ubicacion.objects.select_related("sector").order_by("sector__sector", "ubicacion")

    return render(
        request,
        "home.html",
        {
            "sectores": sectores,
            "ubicaciones": ubicaciones,
            "vista_actual": vista_actual,
        },
    )



# -------------------
# UTIL: DELETE CONFIRM
# -------------------

def _confirm_delete(request, obj, cancel_url_name, cancel_kwargs, success_url_name):
    if request.method == "POST":
        obj.delete()
        return redirect(success_url_name)

    cancel_url = reverse(cancel_url_name, kwargs=cancel_kwargs)
    return render(request, "confirm_delete.html", {"obj": obj, "cancel_url": cancel_url})


# -------------------
# SECTOR
# -------------------

@login_required
def lista_sectores(request):
    sectores = Sector.objects.all().order_by("sector")
    return render(request, "sector/sectores.html", {"sectores": sectores})


@login_required
def detalle_sector(request, sector_id):
    sector = get_object_or_404(Sector, pk=sector_id)
    ubicaciones = Ubicacion.objects.filter(sector=sector).order_by("ubicacion")
    return render(
        request,
        "sector/detalle_sector.html",
        {"sector": sector, "ubicaciones": ubicaciones},
    )


@login_required
def crear_sector(request):
    if request.method == "GET":
        return render(request, "sector/crear_sector.html", {"form": CrearSector()})
    form = CrearSector(request.POST)
    if form.is_valid():
        form.save()
        return redirect("lista_sectores")
    return render(request, "sector/crear_sector.html", {"form": form})


@login_required
def editar_sector(request, sector_id):
    sector = get_object_or_404(Sector, pk=sector_id)
    if request.method == "GET":
        return render(
            request,
            "sector/editar_sector.html",
            {"form": EditarSector(instance=sector), "sector": sector},
        )

    form = EditarSector(request.POST, instance=sector)
    if form.is_valid():
        form.save()
        return redirect("detalle_sector", sector_id=sector.id)
    return render(
        request,
        "sector/editar_sector.html",
        {"form": form, "sector": sector},
    )


@login_required
def borrar_sector(request, sector_id):
    sector = get_object_or_404(Sector, pk=sector_id)
    return _confirm_delete(
        request, sector, "detalle_sector", {"sector_id": sector.id}, "lista_sectores"
    )


# -------------------
# UBICACION
# -------------------

@login_required
def lista_ubicaciones(request):
    # id de sector recibido por GET ?sector=3
    sector_id = request.GET.get("sector", "").strip()

    # query base
    ubicaciones_qs = Ubicacion.objects.select_related("sector").all()

    # si viene sector, filtramos
    if sector_id:
        try:
            ubicaciones_qs = ubicaciones_qs.filter(sector_id=int(sector_id))
        except ValueError:
            # si viene algo raro, ignoramos el filtro
            sector_id = ""

    ubicaciones = ubicaciones_qs.order_by("ubicacion")

    # para armar el combo
    sectores = Sector.objects.all().order_by("sector")

    return render(
        request,
        "ubicacion/ubicaciones.html",
        {
            "ubicaciones": ubicaciones,
            "sectores": sectores,
            "sector_actual": sector_id,  # lo usamos para marcar el option seleccionado
        },
    )



@login_required
def detalle_ubicacion(request, ubicacion_id):
    ubicacion = get_object_or_404(Ubicacion, pk=ubicacion_id)
    pisos = Piso.objects.filter(ubicacion=ubicacion).order_by("piso")
    return render(
        request,
        "ubicacion/detalle_ubicacion.html",
        {"ubicacion": ubicacion, "pisos": pisos},
    )


@login_required
def crear_ubicacion(request):
    if request.method == "GET":
        return render(request, "ubicacion/crear_ubicacion.html", {"form": CrearUbicacion()})
    form = CrearUbicacion(request.POST)
    if form.is_valid():
        ubicacion = form.save()
        return redirect("detalle_sector", sector_id=ubicacion.sector_id)
    return render(request, "ubicacion/crear_ubicacion.html", {"form": form})


@login_required
def editar_ubicacion(request, ubicacion_id):
    ubicacion = get_object_or_404(Ubicacion, pk=ubicacion_id)
    if request.method == "GET":
        return render(
            request,
            "ubicacion/editar_ubicacion.html",
            {"form": EditarUbicacion(instance=ubicacion), "ubicacion": ubicacion},
        )

    form = EditarUbicacion(request.POST, instance=ubicacion)
    if form.is_valid():
        form.save()
        return redirect("detalle_ubicacion", ubicacion_id=ubicacion.id)
    return render(
        request,
        "ubicacion/editar_ubicacion.html",
        {"form": form, "ubicacion": ubicacion},
    )


@login_required
def borrar_ubicacion(request, ubicacion_id):
    ubicacion = get_object_or_404(Ubicacion, pk=ubicacion_id)
    return _confirm_delete(
        request,
        ubicacion,
        "detalle_ubicacion",
        {"ubicacion_id": ubicacion.id},
        "lista_ubicaciones",
    )


# -------------------
# PISO
# -------------------

@login_required
def lista_pisos(request):
    # ?ubicacion=ID que llega por GET
    ubicacion_id = request.GET.get("ubicacion", "").strip()

    # queryset base
    pisos_qs = Piso.objects.select_related("ubicacion", "ubicacion__sector").all()

    # si viene ubicación, filtramos por esa ubicación
    if ubicacion_id:
        try:
            pisos_qs = pisos_qs.filter(ubicacion_id=int(ubicacion_id))
        except ValueError:
            ubicacion_id = ""  # si viene algo raro, ignoramos el filtro

    pisos = pisos_qs.order_by("ubicacion__ubicacion", "piso")

    # para llenar el combo de ubicaciones
    ubicaciones = Ubicacion.objects.select_related("sector").all().order_by("ubicacion")

    return render(
        request,
        "piso/pisos.html",  # deja aquí la ruta que ya usas
        {
            "pisos": pisos,
            "ubicaciones": ubicaciones,
            "ubicacion_actual": ubicacion_id,
        },
    )



@login_required
def detalle_piso(request, piso_id):
    piso = get_object_or_404(Piso, pk=piso_id)
    lugares = Lugar.objects.filter(piso=piso).order_by("nombre_del_lugar")
    return render(
        request,
        "piso/detalle_piso.html",
        {"piso": piso, "lugares": lugares},
    )


@login_required
def crear_piso(request):
    if request.method == "GET":
        return render(request, "piso/crear_piso.html", {"form": CrearPiso()})
    form = CrearPiso(request.POST)
    if form.is_valid():
        piso = form.save()
        return redirect("detalle_ubicacion", ubicacion_id=piso.ubicacion_id)
    return render(request, "piso/crear_piso.html", {"form": form})


@login_required
def editar_piso(request, piso_id):
    piso = get_object_or_404(Piso, pk=piso_id)
    cancel_url = reverse("detalle_piso", kwargs={"piso_id": piso.id})

    if request.method == "GET":
        return render(
            request,
            "editar_generico.html",
            {
                "titulo": "Editar piso",
                "form": EditarPiso(instance=piso),
                "cancel_url": cancel_url,
            },
        )

    form = EditarPiso(request.POST, instance=piso)
    if form.is_valid():
        form.save()
        return redirect("detalle_piso", piso_id=piso.id)

    return render(
        request,
        "editar_generico.html",
        {
            "titulo": "Editar piso",
            "form": form,
            "cancel_url": cancel_url,
        },
    )


@login_required
def borrar_piso(request, piso_id):
    piso = get_object_or_404(Piso, pk=piso_id)
    return _confirm_delete(
        request, piso, "detalle_piso", {"piso_id": piso.id}, "lista_pisos"
    )


# -------------------
# LUGAR
# -------------------

@login_required
def lista_lugares(request):
    ubicacion_id = request.GET.get("ubicacion", "").strip()
    piso_id = request.GET.get("piso", "").strip()

    # queryset base
    lugares_qs = Lugar.objects.select_related(
        "piso",
        "piso__ubicacion",
        "piso__ubicacion__sector",
    ).all()

    # filtro por ubicación
    if ubicacion_id:
        try:
            u_id = int(ubicacion_id)
            lugares_qs = lugares_qs.filter(piso__ubicacion_id=u_id)
        except ValueError:
            ubicacion_id = ""

    # filtro por piso
    if piso_id:
        try:
            p_id = int(piso_id)
            lugares_qs = lugares_qs.filter(piso_id=p_id)
        except ValueError:
            piso_id = ""

    lugares = lugares_qs.order_by(
        "piso__ubicacion__ubicacion",
        "piso__piso",
        "nombre_del_lugar",
    )

    # combos
    ubicaciones = Ubicacion.objects.select_related("sector").all().order_by("ubicacion")
    # OJO: aquí mandamos **todos** los pisos; el JS se encarga de filtrarlos en el combo
    pisos = Piso.objects.select_related("ubicacion").all().order_by(
        "ubicacion__ubicacion",
        "piso",
    )

    return render(
        request,
        "lugar/lugares.html",
        {
            "lugares": lugares,
            "ubicaciones": ubicaciones,
            "pisos": pisos,
            "ubicacion_actual": ubicacion_id,
            "piso_actual": piso_id,
        },
    )




@login_required
def detalle_lugar(request, lugar_id):
    lugar = get_object_or_404(Lugar, pk=lugar_id)
    objetos = (
        ObjetoLugar.objects.select_related("tipo_de_objeto", "tipo_de_objeto__objeto")
        .filter(lugar=lugar)
        .order_by("-fecha")
    )
    return render(
        request,
        "lugar/detalle_lugar.html",
        {"lugar": lugar, "objetos": objetos},
    )


@login_required
def crear_lugar(request):
    if request.method == "GET":
        return render(request, "lugar/crear_lugar.html", {"form": CrearLugar()})
    form = CrearLugar(request.POST)
    if form.is_valid():
        lugar = form.save()
        return redirect("detalle_piso", piso_id=lugar.piso_id)
    return render(request, "lugar/crear_lugar.html", {"form": form})


@login_required
def editar_lugar(request, lugar_id):
    lugar = get_object_or_404(Lugar, pk=lugar_id)
    cancel_url = reverse("detalle_lugar", kwargs={"lugar_id": lugar.id})

    if request.method == "GET":
        return render(
            request,
            "editar_generico.html",
            {
                "titulo": "Editar lugar",
                "form": EditarLugar(instance=lugar),
                "cancel_url": cancel_url,
            },
        )

    form = EditarLugar(request.POST, instance=lugar)
    if form.is_valid():
        form.save()
        return redirect("detalle_lugar", lugar_id=lugar.id)

    return render(
        request,
        "editar_generico.html",
        {
            "titulo": "Editar lugar",
            "form": form,
            "cancel_url": cancel_url,
        },
    )


@login_required
def borrar_lugar(request, lugar_id):
    lugar = get_object_or_404(Lugar, pk=lugar_id)
    return _confirm_delete(
        request, lugar, "detalle_lugar", {"lugar_id": lugar.id}, "lista_lugares"
    )


# -------------------
# OBJETO DEL LUGAR
# -------------------

@login_required
def lista_objetos_lugar(request):
    lugar_id = request.GET.get("lugar", "").strip()
    objeto_id = request.GET.get("objeto", "").strip()
    tipo_id = request.GET.get("tipo", "").strip()
    estado_val = request.GET.get("estado", "").strip()

    qs = ObjetoLugar.objects.select_related(
        "lugar",
        "lugar__piso",
        "lugar__piso__ubicacion",
        "lugar__piso__ubicacion__sector",
        "tipo_de_objeto",
        "tipo_de_objeto__objeto",
        "tipo_de_objeto__objeto__objeto_categoria",
    ).all()

    # Filtros
    if lugar_id:
        try:
            qs = qs.filter(lugar_id=int(lugar_id))
        except ValueError:
            lugar_id = ""

    if objeto_id:
        try:
            qs = qs.filter(tipo_de_objeto__objeto_id=int(objeto_id))
        except ValueError:
            objeto_id = ""

    if tipo_id:
        try:
            qs = qs.filter(tipo_de_objeto_id=int(tipo_id))
        except ValueError:
            tipo_id = ""

    if estado_val:
        qs = qs.filter(estado=estado_val)

    objetos_lugar = qs.order_by(
        "lugar__piso__ubicacion__ubicacion",
        "lugar__piso__piso",
        "lugar__nombre_del_lugar",
        "tipo_de_objeto__objeto__nombre_del_objeto",
    )

    # Datos para los combos
    lugares = (
        Lugar.objects.select_related(
            "piso",
            "piso__ubicacion",
            "piso__ubicacion__sector",
        )
        .all()
        .order_by(
            "piso__ubicacion__ubicacion",
            "piso__piso",
            "nombre_del_lugar",
        )
    )

    objetos = (
        Objeto.objects.select_related("objeto_categoria")
        .all()
        .order_by("objeto_categoria__nombre_de_categoria", "nombre_del_objeto")
    )

    tipos = (
        TipoObjeto.objects.select_related("objeto", "objeto__objeto_categoria")
        .all()
        .order_by("objeto__nombre_del_objeto", "marca", "material")
    )

    # choices del modelo (por ejemplo [("B", "Bueno"), ...])
    estados = ObjetoLugar.ESTADO

    return render(
        request,
        "objeto_lugar/objetos_lugar.html",  # <-- cambia la ruta si tu HTML está en otro lado
        {
            "objetos": objetos_lugar,
            "lugares": lugares,
            "objetos_catalogo": objetos,
            "tipos": tipos,
            "estados": estados,
            "lugar_actual": lugar_id,
            "objeto_actual": objeto_id,
            "tipo_actual": tipo_id,
            "estado_actual": estado_val,
        },
    )


@login_required
def detalle_objeto_lugar(request, objeto_lugar_id):
    obj = get_object_or_404(ObjetoLugar, pk=objeto_lugar_id)
    historicos = (
        HistoricoObjeto.objects
        .filter(objeto_del_lugar=obj)
        .order_by("-fecha_anterior")
    )
    return render(
        request,
        "objeto_lugar/detalle_objeto_lugar.html",
        {"objeto_lugar": obj, "historicos": historicos},
    )


@login_required
def crear_objeto_lugar(request, lugar_id):
    lugar = get_object_or_404(Lugar, pk=lugar_id)

    if request.method == "GET":
        return render(
            request,
            "objeto_lugar/crear_objeto_lugar.html",
            {"form": CrearObjetoLugar(), "lugar": lugar},
        )

    form = CrearObjetoLugar(request.POST)
    if form.is_valid():
        obj = form.save(commit=False)
        obj.lugar = lugar
        obj.save()
        return redirect("detalle_lugar", lugar_id=lugar.id)

    return render(
        request,
        "objeto_lugar/crear_objeto_lugar.html",
        {"form": form, "lugar": lugar},
    )


@login_required
def editar_objeto_lugar(request, objeto_lugar_id):
    objeto_lugar = get_object_or_404(ObjetoLugar, pk=objeto_lugar_id)
    cancel_url = reverse(
        "detalle_objeto_lugar",
        kwargs={"objeto_lugar_id": objeto_lugar.id},
    )

    if request.method == "GET":
        return render(
            request,
            "editar_generico.html",
            {
                "titulo": "Editar objeto del lugar",
                "form": EditarObjetoLugar(instance=objeto_lugar),
                "cancel_url": cancel_url,
            },
        )

    form = EditarObjetoLugar(request.POST, instance=objeto_lugar)
    if form.is_valid():
        form.save()
        return redirect("detalle_objeto_lugar", objeto_lugar_id=objeto_lugar.id)

    return render(
        request,
        "editar_generico.html",
        {
            "titulo": "Editar objeto del lugar",
            "form": form,
            "cancel_url": cancel_url,
        },
    )


@login_required
def borrar_objeto_lugar(request, objeto_lugar_id):
    obj = get_object_or_404(ObjetoLugar, pk=objeto_lugar_id)
    return _confirm_delete(
        request,
        obj,
        "detalle_lugar",
        {"lugar_id": obj.lugar_id},
        "lista_objetos_lugar",
    )

# -------------------
# TIPO LUGAR
# -------------------

@login_required
def lista_tipos_lugar(request):
    tipos = TipoLugar.objects.all().order_by("tipo_de_lugar")
    return render(request, "tipo_lugar/tipos_lugar.html", {"tipos": tipos})


@login_required
def detalle_tipo_lugar(request, tipo_lugar_id):

    tipo = get_object_or_404(TipoLugar, pk=tipo_lugar_id)

    if request.method == "POST":
        ids = request.POST.getlist("tipicos")

        # normalizamos a ints y evitamos basura
        nuevos_ids = []
        for x in ids:
            try:
                nuevos_ids.append(int(x))
            except (TypeError, ValueError):
                pass

        with transaction.atomic():
            TipoLugarObjetoTipico.objects.filter(tipo_lugar=tipo).delete()
            TipoLugarObjetoTipico.objects.bulk_create(
                [
                    TipoLugarObjetoTipico(
                        tipo_lugar=tipo,
                        tipo_objeto_id=tipo_objeto_id,
                        activo=True,
                        orden=i,
                    )
                    for i, tipo_objeto_id in enumerate(nuevos_ids)
                ]
            )

        return redirect("detalle_tipo_lugar", tipo_lugar_id=tipo.id)

    lugares = Lugar.objects.filter(lugar_tipo_lugar=tipo).order_by("nombre_del_lugar")

    tipicos_qs = (
        TipoLugarObjetoTipico.objects.filter(tipo_lugar=tipo, activo=True)
        .select_related("tipo_objeto__objeto__objeto_categoria")
        .order_by(
            "orden",
            "tipo_objeto__objeto__objeto_categoria__nombre_de_categoria",
            "tipo_objeto__objeto__nombre_del_objeto",
            "tipo_objeto__marca",
            "tipo_objeto__material",
        )
    )
    tipicos_ids = list(tipicos_qs.values_list("tipo_objeto_id", flat=True))

    tipos_objeto = (
        TipoObjeto.objects.select_related("objeto__objeto_categoria")
        .all()
        .order_by(
            "objeto__objeto_categoria__nombre_de_categoria",
            "objeto__nombre_del_objeto",
            "marca",
            "material",
        )
    )

    return render(
        request,
        "tipo_lugar/detalle_tipo_lugar.html",
        {
            "tipo": tipo,
            "lugares": lugares,
            "tipicos": tipicos_qs,
            "tipicos_ids": tipicos_ids,
            "tipos_objeto": tipos_objeto,
        },
    )


@login_required
def crear_tipo_lugar(request):
    if request.method == "GET":
        return render(request, "tipo_lugar/crear_tipo_lugar.html", {"form": CrearTipoLugar()})
    form = CrearTipoLugar(request.POST)
    if form.is_valid():
        form.save()
        return redirect("lista_tipos_lugar")
    return render(request, "tipo_lugar/crear_tipo_lugar.html", {"form": form})


@login_required
def editar_tipo_lugar(request, tipo_lugar_id):
    tipo = get_object_or_404(TipoLugar, pk=tipo_lugar_id)
    cancel_url = reverse("detalle_tipo_lugar", kwargs={"tipo_lugar_id": tipo.id})

    if request.method == "GET":
        return render(
            request,
            "editar_generico.html",
            {
                "titulo": "Editar tipo de lugar",
                "form": EditarTipoLugar(instance=tipo),
                "cancel_url": cancel_url,
            },
        )

    form = EditarTipoLugar(request.POST, instance=tipo)
    if form.is_valid():
        form.save()
        return redirect("detalle_tipo_lugar", tipo_lugar_id=tipo.id)

    return render(
        request,
        "editar_generico.html",
        {
            "titulo": "Editar tipo de lugar",
            "form": form,
            "cancel_url": cancel_url,
        },
    )


@login_required
def borrar_tipo_lugar(request, tipo_lugar_id):
    tipo = get_object_or_404(TipoLugar, pk=tipo_lugar_id)
    return _confirm_delete(
        request,
        tipo,
        "detalle_tipo_lugar",
        {"tipo_lugar_id": tipo.id},
        "lista_tipos_lugar",
    )


# -------------------
# CATEGORIA / OBJETO / TIPO OBJETO
# -------------------

@login_required
def lista_categorias(request):
    categorias = CategoriaObjeto.objects.all().order_by("nombre_de_categoria")
    return render(request, "categoria/categorias.html", {"categorias": categorias})


@login_required
def detalle_categoria(request, categoria_id):
    categoria = get_object_or_404(CategoriaObjeto, pk=categoria_id)
    objetos = Objeto.objects.filter(objeto_categoria_id=categoria_id).order_by("nombre_del_objeto")
    return render(
        request,
        "categoria/detalle_categoria.html",
        {"categoria": categoria, "objetos": objetos},
    )


@login_required
def crear_categoria_objeto(request):
    if request.method == "GET":
        return render(
            request,
            "categoria/crear_categoria_objeto.html",
            {"form": CrearCategoriaObjeto()},
        )
    form = CrearCategoriaObjeto(request.POST)
    if form.is_valid():
        form.save()
        return redirect("lista_categorias")
    return render(request, "categoria/crear_categoria_objeto.html", {"form": form})


@login_required
def editar_categoria(request, categoria_id):
    categoria = get_object_or_404(CategoriaObjeto, pk=categoria_id)
    cancel_url = reverse("detalle_categoria", kwargs={"categoria_id": categoria.id})

    if request.method == "GET":
        return render(
            request,
            "editar_generico.html",
            {
                "titulo": "Editar categoría",
                "form": EditarCategoria(instance=categoria),
                "cancel_url": cancel_url,
            },
        )

    form = EditarCategoria(request.POST, instance=categoria)
    if form.is_valid():
        form.save()
        return redirect("detalle_categoria", categoria_id=categoria.id)

    return render(
        request,
        "editar_generico.html",
        {
            "titulo": "Editar categoría",
            "form": form,
            "cancel_url": cancel_url,
        },
    )


@login_required
def borrar_categoria(request, categoria_id):
    categoria = get_object_or_404(CategoriaObjeto, pk=categoria_id)
    return _confirm_delete(
        request,
        categoria,
        "detalle_categoria",
        {"categoria_id": categoria.id},
        "lista_categorias",
    )


@login_required
def lista_objetos(request):
    # ?categoria=ID que llega por GET
    categoria_id = request.GET.get("categoria", "").strip()

    # queryset base
    objetos_qs = Objeto.objects.select_related("objeto_categoria").all()

    # si viene categoría, filtramos
    if categoria_id:
        try:
            objetos_qs = objetos_qs.filter(objeto_categoria_id=int(categoria_id))
        except ValueError:
            categoria_id = ""  # si viene algo raro, ignoramos el filtro

    objetos = objetos_qs.order_by(
        "objeto_categoria__nombre_de_categoria",
        "nombre_del_objeto",
    )

    # para el combo de categorías
    categorias = CategoriaObjeto.objects.all().order_by("nombre_de_categoria")

    return render(
        request,
        "objeto/objetos.html",   # deja aquí la ruta real de tu template
        {
            "objetos": objetos,
            "categorias": categorias,
            "categoria_actual": categoria_id,
        },
    )


@login_required
def detalle_objeto(request, objeto_id):
    objeto = get_object_or_404(Objeto, pk=objeto_id)
    tipos = TipoObjeto.objects.filter(objeto=objeto).order_by("marca", "material")
    return render(
        request,
        "objeto/detalle_objeto.html",
        {"objeto": objeto, "tipos": tipos},
    )


@login_required
def crear_objeto(request):
    if request.method == "GET":
        return render(request, "objeto/crear_objeto.html", {"form": CrearObjeto()})
    form = CrearObjeto(request.POST)
    if form.is_valid():
        form.save()
        return redirect("lista_objetos")
    return render(request, "objeto/crear_objeto.html", {"form": form})


@login_required
def editar_objeto(request, objeto_id):
    objeto = get_object_or_404(Objeto, pk=objeto_id)
    cancel_url = reverse("detalle_objeto", kwargs={"objeto_id": objeto.id})

    if request.method == "GET":
        return render(
            request,
            "editar_generico.html",
            {
                "titulo": "Editar objeto",
                "form": EditarObjeto(instance=objeto),
                "cancel_url": cancel_url,
            },
        )

    form = EditarObjeto(request.POST, instance=objeto)
    if form.is_valid():
        form.save()
        return redirect("detalle_objeto", objeto_id=objeto.id)

    return render(
        request,
        "editar_generico.html",
        {
            "titulo": "Editar objeto",
            "form": form,
            "cancel_url": cancel_url,
        },
    )


@login_required
def borrar_objeto(request, objeto_id):
    objeto = get_object_or_404(Objeto, pk=objeto_id)
    return _confirm_delete(
        request,
        objeto,
        "detalle_objeto",
        {"objeto_id": objeto.id},
        "lista_objetos",
    )


@login_required
def lista_tipos_objeto(request):
    # ?categoria=ID & ?objeto=ID
    categoria_id = request.GET.get("categoria", "").strip()
    objeto_id = request.GET.get("objeto", "").strip()

    # queryset base
    tipos_qs = TipoObjeto.objects.select_related(
        "objeto",
        "objeto__objeto_categoria",
    ).all()

    # filtro por categoría (a través del objeto)
    if categoria_id:
        try:
            c_id = int(categoria_id)
            tipos_qs = tipos_qs.filter(objeto__objeto_categoria_id=c_id)
        except ValueError:
            categoria_id = ""

    # filtro por objeto
    if objeto_id:
        try:
            o_id = int(objeto_id)
            tipos_qs = tipos_qs.filter(objeto_id=o_id)
        except ValueError:
            objeto_id = ""

    tipos_objeto = tipos_qs.order_by(
        "objeto__objeto_categoria__nombre_de_categoria",
        "objeto__nombre_del_objeto",
        "marca",
        "material",
    )

    # ===== combos =====
    categorias = CategoriaObjeto.objects.all().order_by("nombre_de_categoria")

    # mandamos TODOS los objetos, el JS se encarga de mostrar sólo los de la categoría elegida
    objetos = Objeto.objects.select_related("objeto_categoria").all().order_by(
        "objeto_categoria__nombre_de_categoria",
        "nombre_del_objeto",
    )

    return render(
        request,
        "tipo_objeto/tipos_objeto.html",  # ajusta la ruta si tu template está en otro lado
        {
            "tipos_objeto": tipos_objeto,
            "categorias": categorias,
            "objetos": objetos,
            "categoria_actual": categoria_id,
            "objeto_actual": objeto_id,
        },
    )



@login_required
def detalle_tipo_objeto(request, tipo_objeto_id):
    tipo = get_object_or_404(TipoObjeto, pk=tipo_objeto_id)
    usados = (
        ObjetoLugar.objects.filter(tipo_de_objeto=tipo)
        .select_related("lugar")
        .order_by("-fecha")
    )
    return render(
        request,
        "tipo_objeto/detalle_tipo_objeto.html",
        {"tipo": tipo, "usados": usados},
    )


@login_required
def crear_tipo_objeto(request):
    if request.method == "GET":
        return render(request, "tipo_objeto/crear_tipo_objeto.html", {"form": CrearTipoObjeto()})
    form = CrearTipoObjeto(request.POST)
    if form.is_valid():
        form.save()
        return redirect("lista_tipos_objeto")
    return render(request, "tipo_objeto/crear_tipo_objeto.html", {"form": form})


@login_required
def editar_tipo_objeto(request, tipo_objeto_id):
    tipo = get_object_or_404(TipoObjeto, pk=tipo_objeto_id)
    cancel_url = reverse("detalle_tipo_objeto", kwargs={"tipo_objeto_id": tipo.id})

    if request.method == "GET":
        return render(
            request,
            "editar_generico.html",
            {
                "titulo": "Editar tipo de objeto",
                "form": EditarTipoObjeto(instance=tipo),
                "cancel_url": cancel_url,
            },
        )

    form = EditarTipoObjeto(request.POST, instance=tipo)
    if form.is_valid():
        form.save()
        return redirect("detalle_tipo_objeto", tipo_objeto_id=tipo.id)

    return render(
        request,
        "editar_generico.html",
        {
            "titulo": "Editar tipo de objeto",
            "form": form,
            "cancel_url": cancel_url,
        },
    )


@login_required
def borrar_tipo_objeto(request, tipo_objeto_id):
    tipo = get_object_or_404(TipoObjeto, pk=tipo_objeto_id)
    return _confirm_delete(
        request,
        tipo,
        "detalle_tipo_objeto",
        {"tipo_objeto_id": tipo.id},
        "lista_tipos_objeto",
    )


# -------------------
# HISTORICO
# -------------------

@login_required
def lista_historicos(request):
    lugar_id = request.GET.get("lugar", "").strip()
    objeto_id = request.GET.get("objeto", "").strip()
    tipo_id = request.GET.get("tipo", "").strip()
    estado_val = request.GET.get("estado", "").strip()

    qs = HistoricoObjeto.objects.select_related(
        "objeto_del_lugar",
        "objeto_del_lugar__lugar",
        "objeto_del_lugar__lugar__piso",
        "objeto_del_lugar__lugar__piso__ubicacion",
        "objeto_del_lugar__lugar__piso__ubicacion__sector",
        "objeto_del_lugar__tipo_de_objeto",
        "objeto_del_lugar__tipo_de_objeto__objeto",
        "objeto_del_lugar__tipo_de_objeto__objeto__objeto_categoria",
    ).all()

    # ---- Filtros ----
    if lugar_id:
        try:
            qs = qs.filter(objeto_del_lugar__lugar_id=int(lugar_id))
        except ValueError:
            lugar_id = ""

    if objeto_id:
        try:
            qs = qs.filter(objeto_del_lugar__tipo_de_objeto__objeto_id=int(objeto_id))
        except ValueError:
            objeto_id = ""

    if tipo_id:
        try:
            qs = qs.filter(objeto_del_lugar__tipo_de_objeto_id=int(tipo_id))
        except ValueError:
            tipo_id = ""

    if estado_val:
        qs = qs.filter(estado_anterior=estado_val)

    historicos = qs.order_by(
        "-fecha_anterior",
        "objeto_del_lugar__lugar__piso__ubicacion__ubicacion",
        "objeto_del_lugar__lugar__piso__piso",
        "objeto_del_lugar__lugar__nombre_del_lugar",
    )

    # ---- datos para los combos ----
    lugares = (
        Lugar.objects.select_related(
            "piso",
            "piso__ubicacion",
            "piso__ubicacion__sector",
        )
        .all()
        .order_by(
            "piso__ubicacion__ubicacion",
            "piso__piso",
            "nombre_del_lugar",
        )
    )

    objetos = (
        Objeto.objects.select_related("objeto_categoria")
        .all()
        .order_by("objeto_categoria__nombre_de_categoria", "nombre_del_objeto")
    )

    tipos = (
        TipoObjeto.objects.select_related("objeto", "objeto__objeto_categoria")
        .all()
        .order_by("objeto__nombre_del_objeto", "marca", "material")
    )

    # choices del campo estado_anterior
    estados = HistoricoObjeto._meta.get_field("estado_anterior").choices

    return render(
        request,
        "historico/historicos.html",   # pon aquí la ruta real de tu template
        {
            "historicos": historicos,
            "lugares": lugares,
            "objetos_catalogo": objetos,
            "tipos": tipos,
            "estados": estados,
            "lugar_actual": lugar_id,
            "objeto_actual": objeto_id,
            "tipo_actual": tipo_id,
            "estado_actual": estado_val,
        },
    )


@login_required
def detalle_historico(request, historico_id):
    historico = get_object_or_404(HistoricoObjeto, pk=historico_id)
    return render(request, "historico/detalle_historico.html", {"historico": historico})


@login_required
def crear_historico(request, objeto_lugar_id):
    objeto_lugar = get_object_or_404(ObjetoLugar, pk=objeto_lugar_id)

    if request.method == "GET":
        form = CrearHistorico(
            initial={
                "cantidad_anterior": objeto_lugar.cantidad,
                "estado_anterior": objeto_lugar.estado,
                "detalle_anterior": objeto_lugar.detalle,
                "fecha_anterior": objeto_lugar.fecha,
                "objeto_del_lugar": objeto_lugar.id,
            }
        )
        return render(
            request,
            "historico/crear_historico.html",
            {"form": form, "objeto_lugar": objeto_lugar},
        )

    form = CrearHistorico(request.POST)
    if form.is_valid():
        h = form.save(commit=False)
        h.objeto_del_lugar = objeto_lugar
        h.save()
        return redirect("detalle_objeto_lugar", objeto_lugar_id=objeto_lugar.id)

    return render(
        request,
        "historico/crear_historico.html",
        {"form": form, "objeto_lugar": objeto_lugar},
    )


@login_required
def editar_historico(request, historico_id):
    historico = get_object_or_404(HistoricoObjeto, pk=historico_id)
    cancel_url = reverse("detalle_historico", kwargs={"historico_id": historico.id})

    if request.method == "GET":
        return render(
            request,
            "editar_generico.html",
            {
                "titulo": "Editar histórico",
                "form": EditarHistorico(instance=historico),
                "cancel_url": cancel_url,
            },
        )

    form = EditarHistorico(request.POST, instance=historico)
    if form.is_valid():
        form.save()
        return redirect("detalle_historico", historico_id=historico.id)

    return render(
        request,
        "editar_generico.html",
        {
            "titulo": "Editar histórico",
            "form": form,
            "cancel_url": cancel_url,
        },
    )


@login_required
def borrar_historico(request, historico_id):
    historico = get_object_or_404(HistoricoObjeto, pk=historico_id)
    return _confirm_delete(
        request,
        historico,
        "detalle_historico",
        {"historico_id": historico.id},
        "lista_historicos",
    )


# -------------------
# RESUMEN
# -------------------

def _add_percentages(rows):
    for r in rows:
        total = r.get("total") or 0
        b = r.get("buenas") or 0
        p = r.get("pendientes") or 0
        m = r.get("malas") or 0

        if total == 0:
            r["pct_buenas"] = r["pct_pendientes"] = r["pct_malas"] = 0
        else:
            r["pct_buenas"] = round(b * 100 / total, 1)
            r["pct_pendientes"] = round(p * 100 / total, 1)
            r["pct_malas"] = round(m * 100 / total, 1)
    return rows


def resumen_general(request):
    # ----------------------
    # 1) Leer filtros del GET
    # ----------------------
    sector_id = request.GET.get("sector") or None
    ubicacion_id = request.GET.get("ubicacion") or None
    piso_id = request.GET.get("piso") or None
    tipo_lugar_id = request.GET.get("tipo_lugar") or None
    categoria_id = request.GET.get("categoria") or None
    objeto_id = request.GET.get("objeto") or None
    tipo_objeto_id = request.GET.get("tipo_objeto") or None
    estado = request.GET.get("estado") or None
    marca = request.GET.get("marca") or None
    material = request.GET.get("material") or None

    # ----------------------
    # 2) Base de datos filtrada
    # ----------------------
    base_qs = ObjetoLugar.objects.select_related(
        "lugar",
        "lugar__piso",
        "lugar__piso__ubicacion",
        "lugar__piso__ubicacion__sector",
        "lugar__lugar_tipo_lugar",
        "tipo_de_objeto",
        "tipo_de_objeto__objeto",
        "tipo_de_objeto__objeto__objeto_categoria",
    )

    if sector_id:
        base_qs = base_qs.filter(lugar__piso__ubicacion__sector_id=sector_id)
    if ubicacion_id:
        base_qs = base_qs.filter(lugar__piso__ubicacion_id=ubicacion_id)
    if piso_id:
        base_qs = base_qs.filter(lugar__piso_id=piso_id)
    if tipo_lugar_id:
        base_qs = base_qs.filter(lugar__lugar_tipo_lugar_id=tipo_lugar_id)
    if categoria_id:
        base_qs = base_qs.filter(
            tipo_de_objeto__objeto__objeto_categoria_id=categoria_id
        )
    if objeto_id:
        base_qs = base_qs.filter(tipo_de_objeto__objeto_id=objeto_id)
    if tipo_objeto_id:
        base_qs = base_qs.filter(tipo_de_objeto_id=tipo_objeto_id)
    if estado:
        base_qs = base_qs.filter(estado=estado)
    if marca:
        base_qs = base_qs.filter(tipo_de_objeto__marca=marca)
    if material:
        base_qs = base_qs.filter(tipo_de_objeto__material=material)

    # ----------------------
    # 3) Resumen por sector
    # ----------------------
    qs_sector = (
        base_qs.values(
            "lugar__piso__ubicacion__sector__id",
            "lugar__piso__ubicacion__sector__sector",
        )
        .annotate(
            total=Sum("cantidad"),
            buenas=Sum("cantidad", filter=Q(estado="B")),
            pendientes=Sum("cantidad", filter=Q(estado="P")),
            malas=Sum("cantidad", filter=Q(estado="M")),
        )
        .order_by("lugar__piso__ubicacion__sector__sector")
    )
    resumen_sector = list(qs_sector)
    _add_percentages(resumen_sector)

    # ----------------------
    # 4) Resumen por ubicación
    # ----------------------
    qs_ubic = (
        base_qs.values(
            "lugar__piso__ubicacion__id",
            "lugar__piso__ubicacion__ubicacion",
            "lugar__piso__ubicacion__sector__sector",
        )
        .annotate(
            total=Sum("cantidad"),
            buenas=Sum("cantidad", filter=Q(estado="B")),
            pendientes=Sum("cantidad", filter=Q(estado="P")),
            malas=Sum("cantidad", filter=Q(estado="M")),
        )
        .order_by(
            "lugar__piso__ubicacion__sector__sector",
            "lugar__piso__ubicacion__ubicacion",
        )
    )
    resumen_ubic = list(qs_ubic)
    _add_percentages(resumen_ubic)

    # ----------------------
    # 5) Resumen por objeto (objeto, no tipo)
    # ----------------------
    qs_obj = (
        base_qs.values(
            "tipo_de_objeto__objeto__id",
            "tipo_de_objeto__objeto__nombre_del_objeto",
        )
        .annotate(
            total=Sum("cantidad"),
            buenas=Sum("cantidad", filter=Q(estado="B")),
            pendientes=Sum("cantidad", filter=Q(estado="P")),
            malas=Sum("cantidad", filter=Q(estado="M")),
        )
        .order_by("tipo_de_objeto__objeto__nombre_del_objeto")
    )
    resumen_obj = list(qs_obj)
    _add_percentages(resumen_obj)

    # Objetos en estado malo, para el detalle por objeto
    malos_qs = (
        base_qs.filter(estado="M")
        .select_related(
            "lugar__piso__ubicacion__sector",
            "lugar__piso__ubicacion",
            "lugar__piso",
            "lugar",
            "tipo_de_objeto__objeto",
        )
        .order_by(
            "tipo_de_objeto__objeto__nombre_del_objeto",
            "lugar__piso__ubicacion__sector__sector",
            "lugar__piso__ubicacion__ubicacion",
            "lugar__piso__piso",
            "lugar__nombre_del_lugar",
        )
    )

    malos_por_objeto = {}
    for ol in malos_qs:
        oid = ol.tipo_de_objeto.objeto_id
        malos_por_objeto.setdefault(oid, []).append(ol)

    resumen_objetos = []
    for r in resumen_obj:
        oid = r["tipo_de_objeto__objeto__id"]
        resumen_objetos.append(
            {
                "id": oid,
                "nombre": r["tipo_de_objeto__objeto__nombre_del_objeto"],
                "total": r["total"] or 0,
                "buenas": r["buenas"] or 0,
                "pendientes": r["pendientes"] or 0,
                "malas": r["malas"] or 0,
                "pct_buenas": r["pct_buenas"],
                "pct_pendientes": r["pct_pendientes"],
                "pct_malas": r["pct_malas"],
                "malos": malos_por_objeto.get(oid, []),
            }
        )

    # ----------------------
    # 6) Datos para los combos de filtros
    # ----------------------
    sectores = Sector.objects.order_by("sector")
    ubicaciones = Ubicacion.objects.select_related("sector").order_by(
        "sector__sector", "ubicacion"
    )
    pisos = Piso.objects.select_related("ubicacion").order_by(
        "ubicacion__ubicacion", "piso"
    )
    tipos_lugar = TipoLugar.objects.order_by("tipo_de_lugar")
    categorias = CategoriaObjeto.objects.order_by("nombre_de_categoria")
    objetos_catalogo = Objeto.objects.select_related("objeto_categoria").order_by(
        "nombre_del_objeto"
    )
    tipos_objeto = TipoObjeto.objects.select_related("objeto").order_by(
        "objeto__nombre_del_objeto", "marca", "material"
    )

    # NUEVO: marcas y materiales únicos
    marcas = (
        TipoObjeto.objects.exclude(Q(marca__isnull=True) | Q(material__exact="")).values_list("marca", flat=True).distinct().order_by("marca")
    )
    materiales = (
        TipoObjeto.objects.exclude(Q(material__isnull=True) | Q(material__exact="")).values_list("material", flat=True).distinct().order_by("material")
    )

    estados = ObjetoLugar.ESTADO

    contexto = {
        "resumen_sector": resumen_sector,
        "resumen_ubic": resumen_ubic,
        "resumen_objetos": resumen_objetos,
        # combos
        "sectores": sectores,
        "ubicaciones": ubicaciones,
        "pisos": pisos,
        "tipos_lugar": tipos_lugar,
        "categorias": categorias,
        "objetos_catalogo": objetos_catalogo,
        "tipos_objeto": tipos_objeto,
        "estados": estados,
        "marcas": marcas,
        "materiales": materiales,
        # valores seleccionados
        "sector_actual": sector_id or "",
        "ubicacion_actual": ubicacion_id or "",
        "piso_actual": piso_id or "",
        "tipo_lugar_actual": tipo_lugar_id or "",
        "categoria_actual": categoria_id or "",
        "objeto_actual": objeto_id or "",
        "tipo_objeto_actual": tipo_objeto_id or "",
        "estado_actual": estado or "",
        "marca_actual": marca or "",
        "material_actual": material or "",
    }

    return render(request, "resumen/resumen_general.html", contexto)


    # -------------------------
# AJAX: combos dependientes
# -------------------------

def ajax_ubicaciones_por_sector(request):
    """
    Devuelve las ubicaciones ligadas a un sector (para el combo de Ubicación).
    GET: ?sector_id=<id>
    """
    sector_id = request.GET.get("sector_id")
    from .models import Ubicacion  # import local para no romper nada

    qs = Ubicacion.objects.filter(sector_id=sector_id).order_by("ubicacion")
    data = [{"id": u.id, "nombre": u.ubicacion} for u in qs]
    return JsonResponse(data, safe=False)


def ajax_pisos_por_ubicacion(request):
    """
    Devuelve los pisos ligados a una ubicación (para el combo de Piso).
    GET: ?ubicacion_id=<id>
    """
    ubicacion_id = request.GET.get("ubicacion_id")
    from .models import Piso

    qs = Piso.objects.filter(ubicacion_id=ubicacion_id).order_by("piso")
    data = [{"id": p.id, "nombre": f"Piso {p.piso}"} for p in qs]
    return JsonResponse(data, safe=False)


def ajax_lugares_por_piso(request):
    """
    (Por si lo necesitas) Devuelve lugares ligados a un piso.
    GET: ?piso_id=<id>
    """
    piso_id = (request.GET.get("piso_id") or "").strip()
    tipo_lugar_id=(request.GET.get("tipo_lugar_id")or"").strip()
    qs = Lugar.objects.all().order_by("nombre_del_lugar")
    if piso_id.isdigit():
        qs = qs.filter(piso_id=int(piso_id))
    else:
        qs=qs.none()
    

    if tipo_lugar_id.isdigit():
        qs = qs.filter(lugar_tipo_lugar_id=int(tipo_lugar_id))
    
    data = [{"id":l.id, "nombre":l.nombre_del_lugar}for l in qs]
    return JsonResponse(data, safe=False)


def ajax_objetos_por_categoria(request):
    """
    Devuelve los OBJETOS de una categoría (para el combo de Objeto).
    GET: ?categoria_id=<id>
    """
    categoria_id = request.GET.get("categoria_id")
    from .models import Objeto

    qs = Objeto.objects.filter(objeto_categoria_id=categoria_id).order_by(
        "nombre_del_objeto"
    )
    data = [{"id": o.id, "nombre": o.nombre_del_objeto} for o in qs]
    return JsonResponse(data, safe=False)


def ajax_tipos_por_objeto(request):
    """
    Devuelve los TIPOS de objeto (marca/material) ligados a un objeto.
    GET: ?objeto_id=<id>
    """
    objeto_id = request.GET.get("objeto_id")
    from .models import TipoObjeto

    qs = TipoObjeto.objects.filter(objeto_id=objeto_id).order_by("marca", "material")
    data = [
        {
            "id": t.id,
            "nombre": f"{t.marca} {t.material}",
        }
        for t in qs
    ]
    return JsonResponse(data, safe=False)

TIPICOS_POR_TIPO_LUGAR = {
    "Baño": {
        "Infraestructura": [
            "Paredes", "Piso", "Cielo", "Techo", "Luces" ,"Ventanas", "Puertas",
            "Conexión eléctrica", "Interruptores"
        ],
        "Sanitario": [
            "Tasas", "Urinario","Desagües","Papeleros","Lavamanos"
        ],
        "Decoración":[
            "Espejos",
        ],
        "Higiene":[
            "Jaboneras","Dispensadores de papel", "Dispensadores de jabón"
        ],
    },
    "Vestidor": {
        "Infraestructura": [
            "Paredes", "Piso", "Cielo", "Techo", "Luces" ,"Ventanas", "Puertas",
            "Conexión eléctrica", "Interruptores"
        ],
        "Mobiliario": [
            "Bancos", "Casilleros", "Percheros",
        ],
        "Sanitario":[
            "Duchas",
        ],
        "Higiene":[
            "Secadores de toalla","Dispensadores de jabón"
        ],
        "Climatización":[
            "Extractores","Estufas"
        ],
    },
    "Comedor": {
        "Infraestructura": [
            "Paredes", "Piso", "Cielo", "Techo", "Luces" ,"Ventanas", "Puertas",
            "Conexión eléctrica", "Interruptores"
        ],
        "Mobiliario":[
            "Mesas","Sillas","Muebles"
        ],
        "Electrodomésticos": [
            "Refrigerador","Microondas","Dispensador de agua","Televisor",
        ],
        "Sanitario":[
            "Lavaplatos","Papeleros"
        ],
        "Climatización":[
            "Aire acondicionado"
        ],
    },
    "Cafetería":{
        "Infraestructura": [
            "Paredes", "Piso", "Cielo", "Techo", "Luces" ,"Ventanas", "Puertas",
            "Conexión eléctrica", "Interruptores"
        ],
        "Mobiliario":[
            "Mesas","Sillas","Muebles"
        ],
        "Electrodomésticos":[
            "Cafetera","Refrigerador","Dispensador de agua"
        ],
        "Climatización":[
            "Aire acondicionado"
        ],
    },
    "Baño vestidor":{
        "Infraestructura": [
            "Paredes", "Piso", "Cielo", "Techo", "Luces" ,"Ventanas", "Puertas",
            "Conexión eléctrica", "Interruptores"
        ],
        "Sanitario": [
            "Tasas", "Urinario","Desagües","Lavamanos","Duchas","Papeleros"
        ],
        "Decoración":[
            "Espejos",
        ],
        "Higiene":[
            "Secadores de toalla", "Dispensadores de jabón"
        ],
        "Mobiliario":[
            "Bancas","Casilleros",
        ],
        "Climatización":[
            "Extractores"
        ],
    }
}


@require_GET
def objetos_tipicos_por_tipo_lugar(request, tipo_lugar_pk):

    tipo_lugar = get_object_or_404(TipoLugar, pk=tipo_lugar_pk)

    qs = (
        TipoLugarObjetoTipico.objects.filter(tipo_lugar=tipo_lugar, activo=True)
        .select_related("tipo_objeto__objeto__objeto_categoria")
        .order_by(
            "orden",
            "tipo_objeto__objeto__objeto_categoria__nombre_de_categoria",
            "tipo_objeto__objeto__nombre_del_objeto",
            "tipo_objeto__marca",
            "tipo_objeto__material",
        )
    )

    # Seed automático (1 sola vez) desde tu TIPICOS_POR_TIPO_LUGAR, y queda en DB
    if not qs.exists():
        tipicos = TIPICOS_POR_TIPO_LUGAR.get(tipo_lugar.tipo_de_lugar, {})

        if tipicos:
            orden = 0
            with transaction.atomic():
                for nombre_categoria, lista_objetos in tipicos.items():
                    categoria_obj, _ = CategoriaObjeto.objects.get_or_create(
                        nombre_de_categoria=nombre_categoria
                    )

                    for nombre_obj in lista_objetos:
                        obj, _ = Objeto.objects.get_or_create(
                            nombre_del_objeto=nombre_obj,
                            defaults={"objeto_categoria": categoria_obj},
                        )
                        # si ya existía pero con otra categoría, no lo tocamos
                        if obj.objeto_categoria_id != categoria_obj.id:
                            categoria_obj = obj.objeto_categoria

                        tipo_objeto, _ = TipoObjeto.objects.get_or_create(
                            objeto=obj,
                            marca="",
                            material="",
                        )

                        TipoLugarObjetoTipico.objects.get_or_create(
                            tipo_lugar=tipo_lugar,
                            tipo_objeto=tipo_objeto,
                            defaults={"activo": True, "orden": orden},
                        )
                        orden += 1

            qs = (
                TipoLugarObjetoTipico.objects.filter(tipo_lugar=tipo_lugar, activo=True)
                .select_related("tipo_objeto__objeto__objeto_categoria")
                .order_by(
                    "orden",
                    "tipo_objeto__objeto__objeto_categoria__nombre_de_categoria",
                    "tipo_objeto__objeto__nombre_del_objeto",
                    "tipo_objeto__marca",
                    "tipo_objeto__material",
                )
            )

    data = []
    for rel in qs:
        t = rel.tipo_objeto
        cat = t.objeto.objeto_categoria
        obj = t.objeto

        marca = (t.marca or "").strip()
        material = (t.material or "").strip()
        extra = ""
        if marca or material:
            extra = f" ({marca} {material})".strip()

        data.append(
            {
                "categoria_id": cat.id,
                "objeto_id": obj.id,
                "tipo_objeto_id": t.id,
                "label": f"{cat.nombre_de_categoria} - {obj.nombre_del_objeto}{extra}",
            }
        )

    return JsonResponse(data, safe=False)



def _color_por_pct_malas(total, pct_malas):
    if not total:
        return "#9ca3af"  # gris
    if pct_malas >= 30:
        return "#ef4444"  # rojo
    if pct_malas >= 10:
        return "#f59e0b"  # amarillo
    return "#22c55e"      # verde


def _resumen_sector_dict():
    qs = (
        ObjetoLugar.objects
        .values("lugar__piso__ubicacion__sector_id")
        .annotate(
            total=Sum("cantidad"),
            buenas=Sum("cantidad", filter=Q(estado="B")),
            pendientes=Sum("cantidad", filter=Q(estado="P")),
            malas=Sum("cantidad", filter=Q(estado="M")),
        )
    )
    d = {}
    for r in qs:
        total = r["total"] or 0
        malas = r["malas"] or 0
        pendientes = r["pendientes"] or 0
        buenas = r["buenas"] or 0
        pct_malas = round(malas * 100 / total, 1) if total else 0
        pct_pend = round(pendientes * 100 / total, 1) if total else 0
        pct_bue = round(buenas * 100 / total, 1) if total else 0
        d[r["lugar__piso__ubicacion__sector_id"]] = {
            "total": total,
            "pct_malas": pct_malas,
            "pct_pendientes": pct_pend,
            "pct_buenas": pct_bue,
        }
    return d


def _resumen_ubicacion_dict():
    qs = (
        ObjetoLugar.objects
        .values("lugar__piso__ubicacion_id")
        .annotate(
            total=Sum("cantidad"),
            buenas=Sum("cantidad", filter=Q(estado="B")),
            pendientes=Sum("cantidad", filter=Q(estado="P")),
            malas=Sum("cantidad", filter=Q(estado="M")),
        )
    )
    d = {}
    for r in qs:
        total = r["total"] or 0
        malas = r["malas"] or 0
        pendientes = r["pendientes"] or 0
        buenas = r["buenas"] or 0
        pct_malas = round(malas * 100 / total, 1) if total else 0
        pct_pend = round(pendientes * 100 / total, 1) if total else 0
        pct_bue = round(buenas * 100 / total, 1) if total else 0
        d[r["lugar__piso__ubicacion_id"]] = {
            "total": total,
            "pct_malas": pct_malas,
            "pct_pendientes": pct_pend,
            "pct_buenas": pct_bue,
        }
    return d

def _feature(kind, obj, geom, extra_props=None):
    props = {
        "kind": kind,  # "sector" o "ubicacion"
        "id": obj.id,
        "name": obj.sector if kind == "sector" else obj.ubicacion,
        "sector_id": obj.id if kind == "sector" else obj.sector_id,
        "sector_name": obj.sector if kind == "sector" else obj.sector.sector,
        "detail_url": reverse("mapa_sector_detalle", kwargs={"sector_id": obj.id}) if kind == "sector"
                      else reverse("mapa_ubicacion_detalle", kwargs={"ubicacion_id": obj.id}),
        "edit_geom_url": reverse("mapa_sector_editar_geom", kwargs={"sector_id": obj.id}) if kind == "sector"
                         else reverse("mapa_ubicacion_editar_geom", kwargs={"ubicacion_id": obj.id}),
    }
    if extra_props:
        props.update(extra_props)

    return {"type": "Feature", "geometry": geom, "properties": props}




@login_required
def mapa_editor_crear(request):
    sectores = Sector.objects.all().order_by("sector")
    ubicaciones= Ubicacion.objects.select_related("sector").all().order_by("ubicacion")

    return render(request, "mapa/mapa_editor.html", {
        "sectores": sectores,
        "ubicaciones": ubicaciones,
    })


@login_required
def mapa_sector_editar_geom(request, sector_id):
    s = get_object_or_404(Sector, pk=sector_id)
    sectores = Sector.objects.order_by("sector")
    ubicaciones = Ubicacion.objects.select_related("sector").order_by("sector__sector", "ubicacion")

    return render(
        request,
        "mapa/mapa_editor.html",
        {
            "modo": "editar",
            "sectores": sectores,
            "ubicaciones": ubicaciones,
            "editar_tipo": "sector",
            "editar_id": s.id,
            "obj_label": f"Sector: {s.sector}",
            "geom_inicial": s.geom,
        },
    )


@login_required
def mapa_ubicacion_editar_geom(request, ubicacion_id):
    u = get_object_or_404(Ubicacion, pk=ubicacion_id)
    sectores = Sector.objects.order_by("sector")
    ubicaciones = Ubicacion.objects.select_related("sector").order_by("sector__sector", "ubicacion")

    return render(
        request,
        "mapa/mapa_editor.html",
        {
            "modo": "editar",
            "sectores": sectores,
            "ubicaciones": ubicaciones,
            "editar_tipo": "ubicacion",
            "editar_id": u.id,
            "obj_label": f"Ubicación: {u.ubicacion} (Sector {u.sector.sector})",
            "geom_inicial": u.geom,
        },
    )


@login_required
def mapa_sector_detalle(request, sector_id):
    sector = get_object_or_404(Sector, pk=sector_id)
    geom = sector.geom

    # ubicaciones con polígono dentro del sector
    ubic_qs = (
        Ubicacion.objects.select_related("sector")
        .filter(sector_id=sector.id)
        .exclude(geom__isnull=True)
        .order_by("ubicacion")
    )

    # stats por ubicacion (CORRECTO)
    rows = (
        ObjetoLugar.objects
        .filter(lugar__isnull=False, lugar__piso__ubicacion__sector_id=sector.id)
        .values("lugar__piso__ubicacion_id")
        .annotate(
            total=Sum("cantidad"),
            buenas=Sum("cantidad", filter=Q(estado="B")),
            pendientes=Sum("cantidad", filter=Q(estado="P")),
            malas=Sum("cantidad", filter=Q(estado="M")),
        )
    )
    stats = _stats_dict_from_rows(rows, "lugar__piso__ubicacion_id")

    ubic_features = []
    for u in ubic_qs:
        st = stats.get(str(u.id))
        hasPct = bool(st and st["hasPct"])

        ubic_features.append({
            "type": "Feature",
            "geometry": u.geom,
            "properties": {
                "id": u.id,
                "name": u.ubicacion,
                "sector_name": sector.sector,

                "hasPct": hasPct,
                "pct": st["pct_buenas"] if hasPct else 0,

                "total": st["total"] if st else 0,
                "buenas": st["buenas"] if st else 0,
                "pendientes": st["pendientes"] if st else 0,
                "malas": st["malas"] if st else 0,

                "pct_buenas": st["pct_buenas"] if st else 0,
                "pct_pendientes": st["pct_pendientes"] if st else 0,
                "pct_malas": st["pct_malas"] if st else 0,

                "detail_url": reverse("mapa_ubicacion_detalle", kwargs={"ubicacion_id": u.id}),
                "edit_geom_url": reverse("mapa_ubicacion_editar_geom", kwargs={"ubicacion_id": u.id}),
            },
        })

    ubic_fc = {"type": "FeatureCollection", "features": ubic_features}

    return render(request, "mapa/mapa_sector_detalle.html", {
        "sector": sector,
        "geom": geom,
        "ubic_fc": ubic_fc,
    })


def _stats_dict_from_rows(rows, key_field: str):
    out = {}
    for r in rows:
        _id = r[key_field]
        total = int(r.get("total") or 0)
        buenas = int(r.get("buenas") or 0)
        pendientes = int(r.get("pendientes") or 0)
        malas = int(r.get("malas") or 0)

        if total > 0:
            pct_b = round((buenas * 100.0) / total, 1)
            pct_p = round((pendientes * 100.0) / total, 1)
            pct_m = round((malas * 100.0) / total, 1)
            hasPct = True
        else:
            pct_b = pct_p = pct_m = 0.0
            hasPct = False

        out[str(_id)] = {
            "total": total,
            "buenas": buenas,
            "pendientes": pendientes,
            "malas": malas,
            "pct_buenas": pct_b,
            "pct_pendientes": pct_p,
            "pct_malas": pct_m,
            "hasPct": hasPct,
        }
    return out
    
@login_required
def mapa_ubicacion_detalle(request, ubicacion_id):
    ubicacion = get_object_or_404(Ubicacion.objects.select_related("sector"), pk=ubicacion_id)
    geom = ubicacion.geom

    base = ObjetoLugar.objects.filter(
        lugar__isnull=False,
        lugar__piso__ubicacion_id=ubicacion.id
    )

    agg = base.aggregate(
        total=Sum("cantidad"),
        buenas=Sum("cantidad", filter=Q(estado="B")),
        pendientes=Sum("cantidad", filter=Q(estado="P")),
        malas=Sum("cantidad", filter=Q(estado="M")),
    )

    total = int(agg.get("total") or 0)
    buenas = int(agg.get("buenas") or 0)
    pendientes = int(agg.get("pendientes") or 0)
    malas = int(agg.get("malas") or 0)

    if total > 0:
        pct_b = round((buenas * 100.0) / total, 1)
        pct_p = round((pendientes * 100.0) / total, 1)
        pct_m = round((malas * 100.0) / total, 1)
    else:
        pct_b = pct_p = pct_m = 0.0

    resumen = {
        "total": total,
        "buenas": buenas,
        "pendientes": pendientes,
        "malas": malas,
        "pct_buenas": pct_b,
        "pct_pendientes": pct_p,
        "pct_malas": pct_m,
    }

    pisos = Piso.objects.filter(ubicacion_id=ubicacion.id).order_by("piso")

    # stats por piso
    piso_rows = (
        base.values("lugar__piso_id")
        .annotate(
            total=Sum("cantidad"),
            buenas=Sum("cantidad", filter=Q(estado="B")),
            pendientes=Sum("cantidad", filter=Q(estado="P")),
            malas=Sum("cantidad", filter=Q(estado="M")),
        )
    )
    piso_stats = _stats_dict_from_rows(piso_rows, "lugar__piso_id")

    # stats por lugar
    lugar_rows = (
        base.values("lugar_id")
        .annotate(
            total=Sum("cantidad"),
            buenas=Sum("cantidad", filter=Q(estado="B")),
            pendientes=Sum("cantidad", filter=Q(estado="P")),
            malas=Sum("cantidad", filter=Q(estado="M")),
        )
    )
    lugar_stats = _stats_dict_from_rows(lugar_rows, "lugar_id")

    pisos_info = []
    for p in pisos:
        lugares = (
            Lugar.objects
            .select_related("lugar_tipo_lugar")
            .filter(piso_id=p.id)
            .order_by("nombre_del_lugar")
        )

        pstat = piso_stats.get(str(p.id), {
            "total": 0, "buenas": 0, "pendientes": 0, "malas": 0,
            "pct_buenas": 0, "pct_pendientes": 0, "pct_malas": 0,
            "hasPct": False
        })

        lugares_info = []
        for l in lugares:
            lstat = lugar_stats.get(str(l.id), {
                "total": 0, "buenas": 0, "pendientes": 0, "malas": 0,
                "pct_buenas": 0, "pct_pendientes": 0, "pct_malas": 0,
                "hasPct": False
            })
            lugares_info.append({
                "lugar": l,
                "stats": lstat,
                "url_detalle": reverse("detalle_lugar", kwargs={"lugar_id": l.id}),
                "url_lista_objetos": f"{reverse('lista_objetos_lugar')}?lugar={l.id}",
            })

        pisos_info.append({
            "piso": p,
            "stats": pstat,
            "url_detalle": reverse("detalle_piso", kwargs={"piso_id": p.id}),
            "url_lista_lugares": f"{reverse('lista_lugares')}?piso={p.id}",
            "lugares": lugares_info,
        })

    return render(request, "mapa/mapa_ubicacion_detalle.html", {
        "ubicacion": ubicacion,
        "geom": geom,
        "resumen": resumen,
        "pisos_info": pisos_info,
        "url_detalle_ubicacion": reverse("detalle_ubicacion", kwargs={"ubicacion_id": ubicacion.id}),
        "url_lista_pisos": f"{reverse('lista_pisos')}?ubicacion={ubicacion.id}",
        "url_lista_lugares": f"{reverse('lista_lugares')}?ubicacion={ubicacion.id}",
        "url_lista_objetos_lugar": reverse("lista_objetos_lugar"),
    })



@require_POST
@login_required
@transaction.atomic
def mapa_guardar(request):
    """
    Guarda polígono (GeoJSON Polygon) en Sector.geom o Ubicacion.geom.
    - Crear: permite Sector existente/nuevo o Ubicación existente/nueva.
    - Editar: reemplaza geom del objetivo.
    """
    geom_json = (request.POST.get("geom_json") or "").strip()
    if not geom_json:
        return redirect("mapa_admin")

    try:
        geom = json.loads(geom_json)
    except Exception:
        return redirect("mapa_admin")

    modo = request.POST.get("accion", "crear")

    # ========= EDITAR =========
    if modo == "editar":
        editar_tipo = request.POST.get("editar_tipo")
        editar_id = request.POST.get("editar_id")
        try:
            editar_id = int(editar_id)
        except Exception:
            return redirect("mapa_admin")

        if editar_tipo == "sector":
            s = get_object_or_404(Sector, pk=editar_id)
            s.geom = geom
            s.save()
            return redirect("mapa_sector_detalle", sector_id=s.id)

        if editar_tipo == "ubicacion":
            u = get_object_or_404(Ubicacion, pk=editar_id)
            u.geom = geom
            u.save()
            return redirect("mapa_ubicacion_detalle", ubicacion_id=u.id)

        return redirect("mapa_admin")

    # ========= CREAR =========
    tipo_registro = request.POST.get("tipo_registro", "sector")  # "sector" | "ubicacion"

    if tipo_registro == "sector":
        sector_existente = request.POST.get("sector_existente", "").strip()
        sector_nuevo = (request.POST.get("sector_nuevo") or "").strip()

        sector_obj = None
        if sector_existente:
            try:
                sector_obj = Sector.objects.get(pk=int(sector_existente))
            except Exception:
                sector_obj = None

        if not sector_obj and sector_nuevo:
            sector_obj, _ = Sector.objects.get_or_create(sector=sector_nuevo)

        if not sector_obj:
            return redirect("mapa_crear")

        sector_obj.geom = geom
        sector_obj.save()
        return redirect("mapa_sector_detalle", sector_id=sector_obj.id)

    # tipo_registro == "ubicacion"
    sector_id = request.POST.get("sector_para_ubicacion", "").strip()
    try:
        sector_obj = Sector.objects.get(pk=int(sector_id))
    except Exception:
        return redirect("mapa_crear")

    ubicacion_existente = request.POST.get("ubicacion_existente", "").strip()
    ubicacion_nueva = (request.POST.get("ubicacion_nueva") or "").strip()

    ubic_obj = None
    if ubicacion_existente:
        try:
            ubic_obj = Ubicacion.objects.get(pk=int(ubicacion_existente))
        except Exception:
            ubic_obj = None

    if not ubic_obj and ubicacion_nueva:
        ubic_obj, _ = Ubicacion.objects.get_or_create(
            ubicacion=ubicacion_nueva,
            defaults={"sector": sector_obj},
        )
        # si ya existía con otro sector, NO lo cambiamos

    if not ubic_obj:
        return redirect("mapa_crear")

    ubic_obj.geom = geom
    ubic_obj.save()
    return redirect("mapa_ubicacion_detalle", ubicacion_id=ubic_obj.id)


@login_required
def mapa_sector_quitar_geom(request, sector_id):
    s = get_object_or_404(Sector, pk=sector_id)
    if request.method == "POST":
        s.geom = None
        s.save()
        return redirect("mapa_sector_detalle", sector_id=s.id)

    cancel_url = reverse("mapa_sector_detalle", kwargs={"sector_id": s.id})
    return render(request, "mapa/confirm_quitar_geom.html", {"obj": s, "cancel_url": cancel_url})


@login_required
def mapa_ubicacion_quitar_geom(request, ubicacion_id):
    u = get_object_or_404(Ubicacion, pk=ubicacion_id)
    if request.method == "POST":
        u.geom = None
        u.save()
        return redirect("mapa_ubicacion_detalle", ubicacion_id=u.id)

    cancel_url = reverse("mapa_ubicacion_detalle", kwargs={"ubicacion_id": u.id})
    return render(request, "mapa/confirm_quitar_geom.html", {"obj": u, "cancel_url": cancel_url})



def _color_por_pct_buenas(pct):
    if pct is None:
            return "#9ca3af"
    if pct >= 65:
        intensidad = min(max((pct - 65)/ 35, 0),1)
        return  f"rgb({int(34 - 10*intensidad)},{int(197+30*intensidad)},{int(94-40*intensidad)})"
    
    if pct >= 50:
        intensidad = (pct -50)/15
        return f"rgb(245, {int(200+20*intensidad)}, 11)"
    
    intensidad = min(max(pct / 50, 0),1)
    return f"rgb({int(239 + 10*(1-intensidad))}, {int(68 * intensidad)}, {int(68 * intensidad)})"

#################################################################################################################################




def _stats_sector_dict():
    rows = (
        ObjetoLugar.objects
        .filter(lugar__isnull=False)
        .values("lugar__piso__ubicacion__sector_id")
        .annotate(
            total=Sum("cantidad"),
            buenas=Sum(Case(When(estado="B", then=F("cantidad")), default=Value(0), output_field=IntegerField())),
            pendientes=Sum(Case(When(estado="P", then=F("cantidad")), default=Value(0), output_field=IntegerField())),
            malas=Sum(Case(When(estado="M", then=F("cantidad")), default=Value(0), output_field=IntegerField())),
        )
    )

    out = {}
    for r in rows:
        sid = r["lugar__piso__ubicacion__sector_id"]
        total = int(r["total"] or 0)
        buenas = int(r["buenas"] or 0)
        pendientes = int(r["pendientes"] or 0)
        malas = int(r["malas"] or 0)

        if total > 0:
            pct_b = round((buenas * 100.0) / total, 1)
            pct_p = round((pendientes * 100.0) / total, 1)
            pct_m = round((malas * 100.0) / total, 1)
        else:
            pct_b = pct_p = pct_m = 0.0

        out[str(sid)] = {
            "total": total,
            "buenas": buenas,
            "pendientes": pendientes,
            "malas": malas,
            "pct_buenas": pct_b,
            "pct_pendientes": pct_p,
            "pct_malas": pct_m,
        }
    return out


def _stats_ubicacion_dict():
    rows = (
        ObjetoLugar.objects
        .filter(lugar__isnull=False)
        .values("lugar__piso__ubicacion_id")
        .annotate(
            total=Sum("cantidad"),
            buenas=Sum(Case(When(estado="B", then=F("cantidad")), default=Value(0), output_field=IntegerField())),
            pendientes=Sum(Case(When(estado="P", then=F("cantidad")), default=Value(0), output_field=IntegerField())),
            malas=Sum(Case(When(estado="M", then=F("cantidad")), default=Value(0), output_field=IntegerField())),
        )
    )

    out = {}
    for r in rows:
        uid = r["lugar__piso__ubicacion_id"]
        total = int(r["total"] or 0)
        buenas = int(r["buenas"] or 0)
        pendientes = int(r["pendientes"] or 0)
        malas = int(r["malas"] or 0)

        if total > 0:
            pct_b = round((buenas * 100.0) / total, 1)
            pct_p = round((pendientes * 100.0) / total, 1)
            pct_m = round((malas * 100.0) / total, 1)
        else:
            pct_b = pct_p = pct_m = 0.0

        out[str(uid)] = {
            "total": total,
            "buenas": buenas,
            "pendientes": pendientes,
            "malas": malas,
            "pct_buenas": pct_b,
            "pct_pendientes": pct_p,
            "pct_malas": pct_m,
        }
    return out


def construir_geojson_para_mapa():
    features = []

    for s in Sector.objects.all():
        if not s.geom:
            continue
        features.append({
            "type": "Feature",
            "geometry": s.geom,
            "properties": {
                "kind": "sector",
                "id": s.id,
                "sector_id": s.id,
                "name": s.sector,
                "sector_name": s.sector,
                "detail_url": reverse("mapa_sector_detalle", kwargs={"sector_id": s.id}),
                "edit_geom_url": reverse("mapa_sector_editar_geom", kwargs={"sector_id": s.id}),
            },
        })

    for u in Ubicacion.objects.select_related("sector").all():
        if not u.geom:
            continue
        features.append({
            "type": "Feature",
            "geometry": u.geom,
            "properties": {
                "kind": "ubicacion",
                "id": u.id,
                "sector_id": u.sector_id,
                "sector_name": u.sector.sector,
                "name": u.ubicacion,
                "detail_url": reverse("mapa_ubicacion_detalle", kwargs={"ubicacion_id": u.id}),
                "edit_geom_url": reverse("mapa_ubicacion_editar_geom", kwargs={"ubicacion_id": u.id}),
            },
        })

    return {"type": "FeatureCollection", "features": features}


@login_required
def mapa_admin(request):
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return JsonResponse({
            "sector": _stats_sector_dict(),
            "ubicacion": _stats_ubicacion_dict(),
        })

    context = {
        "sectores": Sector.objects.all().order_by("sector"),
        "ubicaciones": Ubicacion.objects.select_related("sector").all().order_by("ubicacion"),
        "mapa_geojson": construir_geojson_para_mapa(),
    }
    return render(request, "mapa/mapa_admin.html", context)


# =========================
# Helpers: parsing Excel
# =========================
def _norm_header(x: object) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s.lower().replace(" ", "").replace("_", "")

def _clean_text(x: object) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    return "" if s == "-" else s

HEADER_ALIASES = {
    "ubicacion": "ubicacion",
    "ubicación": "ubicacion",
    "sector": "sector",
    "piso": "piso",
    "tipodelugar": "tipo_de_lugar",
    "tipolugar": "tipo_de_lugar",
    "tipo_de_lugar": "tipo_de_lugar",
    "lugar": "lugar",
    "categoria": "categoria",
    "categoría": "categoria",
    "objeto": "objeto",
    "tipo": "tipo_objeto",
    "tipoobjeto": "tipo_objeto",
    "tipo_objeto": "tipo_objeto",
    "marca": "marca",
    "material": "material",
    "cantidad": "cantidad",
    "estado": "estado",
    "detalle": "detalle",
    "especificacion": "detalle",
    "especificación": "detalle",
    "fecha": "fecha",
}

# Para reconocer el formato normalizado (tu plantilla)
REQUIRED_HEADERS_NORMALIZADO = {"ubicacion", "sector", "lugar", "objeto", "cantidad", "estado"}

# Para reconocer el formato exportado por TU sistema
RE_SECTOR_UBI = re.compile(r"sector:\s*(.*?)\s*\|\s*ubicaci[oó]n:\s*(.*)", re.IGNORECASE)
RE_PISO = re.compile(r"^\s*piso\s*\d+", re.IGNORECASE)
RE_TIPO_LUGAR = re.compile(r"^\s*tipo\s*de\s*lugar\s*:\s*(.*)$", re.IGNORECASE)

def _row_first_text(ws, r: int) -> str:
    """Primer texto no vacío de una fila (sirve para celdas mergeadas)."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        if v not in (None, ""):
            return str(v).strip()
    return ""

def _find_sector_ubicacion(ws):
    """Busca 'Sector: X | Ubicación: Y' en las primeras filas."""
    for r in range(1, min(ws.max_row, 15) + 1):
        # revisamos varias columnas por si no está en A (merge raro)
        for c in range(1, min(ws.max_column, 8) + 1):
            v = ws.cell(r, c).value
            if not v:
                continue
            m = RE_SECTOR_UBI.search(str(v))
            if m:
                return m.group(1).strip(), m.group(2).strip()
    return None, None


def _parse_normalizado(wb):
    """
    Formato normalizado:
    hoja ObjetosLugar o primera hoja con headers:
    ubicacion, sector, lugar, objeto, cantidad, estado (y opcionales)
    """
    ws = wb["ObjetosLugar"] if "ObjetosLugar" in wb.sheetnames else wb.worksheets[0]

    header_row = None
    header_map = {}

    for r in range(1, min(ws.max_row, 40) + 1):
        mapped = []
        for c in range(1, ws.max_column + 1):
            key = HEADER_ALIASES.get(_norm_header(ws.cell(r, c).value))
            if key:
                mapped.append((c, key))

        keys = set(k for _, k in mapped)
        if REQUIRED_HEADERS_NORMALIZADO.issubset(keys):
            header_row = r
            header_map = {col: key for col, key in mapped}
            break

    if not header_row:
        return []

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        if not any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1)):
            continue

        rec = {}
        for col, key in header_map.items():
            rec[key] = ws.cell(r, col).value

        rows.append({
            "ubicacion": _clean_text(rec.get("ubicacion")),
            "sector": _clean_text(rec.get("sector")),
            "piso": _clean_text(rec.get("piso")),
            "tipo_de_lugar": _clean_text(rec.get("tipo_de_lugar")) or "Sin especificar",
            "lugar": _clean_text(rec.get("lugar")),
            "categoria": _clean_text(rec.get("categoria")) or "Sin categoría",
            "objeto": _clean_text(rec.get("objeto")),
            "tipo_objeto": _clean_text(rec.get("tipo_objeto")),
            "marca": _clean_text(rec.get("marca")),
            "material": _clean_text(rec.get("material")),
            "cantidad": rec.get("cantidad") if rec.get("cantidad") is not None else 0,
            "estado": _clean_text(rec.get("estado")),
            "detalle": _clean_text(rec.get("detalle")),
            "fecha": rec.get("fecha"),
        })

    return rows


def _parse_exportado(wb):
    """
    Formato exportado por tu sistema (NUEVO):
    - "Sector: X | Ubicación: Y"
    - "PISO 1"
    - "Tipo de lugar: Baño"
    - (Fila) Nombre del lugar
    - (Fila) headers en A,C,E,G,I,K,M: Categoría | Objeto | Tipo | Cantidad | Estado | Detalle | Fecha
    - datos abajo
    """
    rows = []

    for ws in wb.worksheets:
        sector, ubicacion = _find_sector_ubicacion(ws)
        if not sector or not ubicacion:
            continue

        current_piso = ""
        current_tipo_lugar = "Sin especificar"

        r = 1
        max_r = ws.max_row

        while r <= max_r:
            txt = _row_first_text(ws, r)

            # PISO
            if txt and RE_PISO.match(txt):
                current_piso = txt # ej "PISO 1"
                r += 1
                continue

            # Tipo de lugar
            if txt:
                mt = RE_TIPO_LUGAR.match(txt)
                if mt:
                    current_tipo_lugar = (mt.group(1) or "").strip() or "Sin especificar"
                    r += 1
                    continue

            # Detectar fila de headers de tabla (buscamos "objeto"+"cantidad"+"estado" en cualquier columna)
            mapped = []
            for c in range(1, ws.max_column + 1):
                key = HEADER_ALIASES.get(_norm_header(ws.cell(r, c).value))
                if key:
                    mapped.append((c, key))
            keys = set(k for _, k in mapped)

            # headers mínimos del bloque
            if {"objeto", "cantidad", "estado"}.issubset(keys):
                header_map = {col: key for col, key in mapped}

                # buscar nombre del lugar hacia arriba (normalmente r-1)
                lugar = ""
                for back in (1, 2, 3, 4):
                    if r - back < 1:
                        break
                    tback = _row_first_text(ws, r - back)
                    if not tback:
                        continue
                    if RE_PISO.match(tback):
                        continue
                    if RE_TIPO_LUGAR.match(tback):
                        continue
                    # esta debería ser la fila mergeada del lugar
                    lugar = tback
                    break

                if not lugar:
                    lugar = "Sin lugar"

                # columnas relevantes (si no están, quedan None)
                col_categoria = next((c for c, k in header_map.items() if k == "categoria"), None)
                col_objeto = next((c for c, k in header_map.items() if k == "objeto"), None)
                col_tipo = next((c for c, k in header_map.items() if k == "tipo_objeto"), None)
                col_cantidad = next((c for c, k in header_map.items() if k == "cantidad"), None)
                col_estado = next((c for c, k in header_map.items() if k == "estado"), None)
                col_detalle = next((c for c, k in header_map.items() if k == "detalle"), None)
                col_fecha = next((c for c, k in header_map.items() if k == "fecha"), None)

                rr = r + 1
                while rr <= max_r:
                    # corte por nuevo bloque
                    tline = _row_first_text(ws, rr)
                    if tline and (RE_PISO.match(tline) or RE_TIPO_LUGAR.match(tline)):
                        break

                    # objeto vacío => fin de tabla
                    obj_val = ws.cell(rr, col_objeto).value if col_objeto else None
                    if obj_val is None or str(obj_val).strip() == "":
                        break

                    cat_val = ws.cell(rr, col_categoria).value if col_categoria else None
                    tipo_val = ws.cell(rr, col_tipo).value if col_tipo else None
                    cant_val = ws.cell(rr, col_cantidad).value if col_cantidad else 0
                    est_val = ws.cell(rr, col_estado).value if col_estado else ""
                    det_val = ws.cell(rr, col_detalle).value if col_detalle else ""
                    fec_val = ws.cell(rr, col_fecha).value if col_fecha else None

                    rows.append({
                        "ubicacion": ubicacion,
                        "sector": sector,
                        "piso": current_piso,
                        "tipo_de_lugar": current_tipo_lugar or "Sin especificar",
                        "lugar": lugar,
                        "categoria": _clean_text(cat_val) or "Sin categoría",
                        "objeto": _clean_text(obj_val),
                        "tipo_objeto": _clean_text(tipo_val),
                        "cantidad": cant_val if cant_val is not None else 0,
                        "estado": _clean_text(est_val),
                        "detalle": _clean_text(det_val),
                        "fecha": fec_val,
                    })
                    rr += 1

                r = rr
                continue

            r += 1

    return rows



def parse_excel(file_obj):
    file_obj.seek(0)
    wb = load_workbook(file_obj, data_only=True)

    rows = _parse_normalizado(wb)
    if rows:
        return rows, "normalizado"

    rows = _parse_exportado(wb)
    if rows:
        return rows, "exportado"

    raise ValueError(
        "No pude reconocer el formato del Excel. "
        "Acepto: (1) plantilla normalizada (con columnas ubicacion/sector/lugar/objeto/cantidad/estado), "
        "o (2) el Excel exportado por tu sistema (Sector|Ubicación, PISO, Tipo de lugar:, LUGAR, y tabla con Categoría/Objeto/Tipo/Cantidad/Estado/Detalle/Fecha)."
    )



def _split_tipo(tipo_str: str):
    """
    Convierte 'Marca - Material' a (marca, material)
    """
    tipo_str = (tipo_str or "").strip()
    if " - " in tipo_str:
        marca, material = [x.strip() for x in tipo_str.split(" - ", 1)]
        return marca, material
    if "-" in tipo_str:
        marca, material = [x.strip() for x in tipo_str.split("-", 1)]
        return marca, material
    return tipo_str, ""


APP_LABEL = Ubicacion._meta.app_label


def _get_model(name: str):
    try:
        return apps.get_model(APP_LABEL, name)
    except LookupError:
        return None


def _field_names(model):
    return {f.name for f in model._meta.get_fields()}


def _pick_field(model, candidates, required=True):
    fields = _field_names(model)
    for c in candidates:
        if c in fields:
            return c
    if not required:
        return None
    raise RuntimeError(f"En {model.__name__} no encontré ninguno de estos campos: {candidates}")


def _pick_fk(model, related_model, candidates=None, required=True):
    # 1) probar por nombre conocido
    if candidates:
        for c in candidates:
            try:
                f = model._meta.get_field(c)
                if getattr(f, "remote_field", None) and f.remote_field.model == related_model:
                    return c
            except Exception:
                pass

    # 2) buscar el primer FK que apunte a related_model
    for f in model._meta.get_fields():
        if getattr(f, "many_to_one", False) and getattr(f, "remote_field", None):
            if f.remote_field.model == related_model:
                return f.name

    if not required:
        return None

    raise RuntimeError(f"No encontré FK desde {model.__name__} hacia {related_model.__name__}")


def import_from_rows(rows):
    """
    Importa filas del payload y crea/actualiza según TU esquema real:
    Sector -> Ubicacion(FK Sector) -> Piso(FK Ubicacion) -> TipoLugar -> Lugar(FK Piso + FK TipoLugar)
    CategoriaObjeto -> Objeto(FK Categoria) -> TipoObjeto -> ObjetoLugar
    """
    import re
    from django.db import transaction

    def _piso_to_int(piso_raw: str) -> int:
        s = (piso_raw or "").strip()
        m = re.search(r"(\d+)", s)
        return int(m.group(1)) if m else 0

    def _estado_to_code(estado_raw: str) -> str:
        s = (estado_raw or "").strip().lower()
        if s in ("b", "bueno"):
            return "B"
        if s in ("p", "pendiente"):
            return "P"
        if s in ("m", "malo"):
            return "M"
        # si viene vacío o raro, por defecto Bueno
        return "B"

    def key(x):
        return (x or "").strip().lower()

    created_ol = 0
    updated_ol = 0

    # caches
    cache_sector = {}
    cache_ubic = {}
    cache_piso = {}
    cache_tl = {}
    cache_lugar = {}
    cache_cat = {}
    cache_obj = {}
    cache_tipo = {}

    with transaction.atomic():
        for r in rows:
            ubicacion = _clean_text(r.get("ubicacion"))
            sector = _clean_text(r.get("sector"))
            piso_raw = _clean_text(r.get("piso"))
            tipo_de_lugar = _clean_text(r.get("tipo_de_lugar")) or "Sin especificar"
            lugar = _clean_text(r.get("lugar"))

            categoria = _clean_text(r.get("categoria")) or "Sin categoría"
            objeto = _clean_text(r.get("objeto"))
            tipo_objeto_str = _clean_text(r.get("tipo_objeto"))
            detalle = _clean_text(r.get("detalle"))
            estado = _estado_to_code(r.get("estado"))

            try:
                cantidad = int(float(r.get("cantidad") or 0))
            except Exception:
                cantidad = 0

            # mínimos
            if not (ubicacion and sector and lugar and objeto):
                continue

            # -------- Sector --------
            ks = key(sector)
            if ks in cache_sector:
                sec_obj = cache_sector[ks]
            else:
                sec_obj, _ = Sector.objects.get_or_create(sector=sector)
                cache_sector[ks] = sec_obj

            # -------- Ubicacion (FK a Sector) --------
            कु = (key(ubicacion), sec_obj.id)
            if कु in cache_ubic:
                ubi_obj = cache_ubic[कु]
            else:
                ubi_obj, _ = Ubicacion.objects.get_or_create(
                    ubicacion=ubicacion,
                    sector=sec_obj,
                )
                cache_ubic[कु] = ubi_obj

            # -------- Piso (FK a Ubicacion) --------
            piso_int = _piso_to_int(piso_raw)
            kp = (piso_int, ubi_obj.id)
            if kp in cache_piso:
                piso_obj = cache_piso[kp]
            else:
                piso_obj, _ = Piso.objects.get_or_create(
                    piso=piso_int,
                    ubicacion=ubi_obj,
                )
                cache_piso[kp] = piso_obj

            # -------- TipoLugar --------
            ktl = key(tipo_de_lugar)
            if ktl in cache_tl:
                tl_obj = cache_tl[ktl]
            else:
                tl_obj, _ = TipoLugar.objects.get_or_create(tipo_de_lugar=tipo_de_lugar)
                cache_tl[ktl] = tl_obj

            # -------- Lugar (FK Piso + TipoLugar) --------
            kl = (key(lugar), piso_obj.id, tl_obj.id)
            if kl in cache_lugar:
                lug_obj = cache_lugar[kl]
            else:
                lug_obj, _ = Lugar.objects.get_or_create(
                    nombre_del_lugar=lugar,
                    piso=piso_obj,
                    lugar_tipo_lugar=tl_obj,
                )
                cache_lugar[kl] = lug_obj

            # -------- CategoriaObjeto --------
            kc = key(categoria)
            if kc in cache_cat:
                cat_obj = cache_cat[kc]
            else:
                cat_obj, _ = CategoriaObjeto.objects.get_or_create(
                    nombre_de_categoria=categoria
                )
                cache_cat[kc] = cat_obj

            # -------- Objeto (FK a Categoria por defecto) --------
            ko = (key(objeto), cat_obj.id)
            if ko in cache_obj:
                obj_obj = cache_obj[ko]
            else:
                obj_obj, _ = Objeto.objects.get_or_create(
                    nombre_del_objeto=objeto,
                    defaults={"objeto_categoria": cat_obj},
                )
                cache_obj[ko] = obj_obj

            # -------- TipoObjeto --------
            marca, material = _split_tipo(tipo_objeto_str)
            kt = (obj_obj.id, key(marca), key(material))
            if kt in cache_tipo:
                tipo_obj = cache_tipo[kt]
            else:
                tipo_obj, _ = TipoObjeto.objects.get_or_create(
                    objeto=obj_obj,
                    marca=marca,
                    material=material,
                )
                cache_tipo[kt] = tipo_obj

            # -------- ObjetoLugar (update_or_create) --------
            ol_obj, created = ObjetoLugar.objects.update_or_create(
                lugar=lug_obj,
                tipo_de_objeto=tipo_obj,
                defaults={
                    "cantidad": cantidad,
                    "estado": estado,
                    "detalle": detalle,
                },
            )

            if created:
                created_ol += 1
            else:
                updated_ol += 1

    return {"created": created_ol, "updated": updated_ol}


# =========================
# Vista
# =========================
@require_http_methods(["GET", "POST"])
def carga_masiva(request):
    # GET -> pantalla upload
    if request.method == "GET":
        return render(request, "excel/carga_masiva.html", {"step": "upload"})

    # POST (guardar)
    if request.POST.get("payload_json"):
        try:
            raw = json.loads(request.POST["payload_json"])
            rows = raw.get("objetos_lugar", [])
            result = import_from_rows(rows)
            messages.success(
                request,
                f"Importación OK. Creados: {result['created']} | Actualizados: {result['updated']}"
            )
            return redirect("carga_masiva")
        except Exception as e:
            return render(
                request,
                "excel/carga_masiva.html",
                {
                    "step": "preview",
                    "error": str(e),
                    "payload": raw if isinstance(raw, dict) else {"objetos_lugar": []},
                    "detected": "manual",
                },
            )

    # POST (subir archivo y previsualizar)
    file_obj = request.FILES.get("archivo")
    if not file_obj:
        return render(
            request,
            "excel/carga_masiva.html",
            {"step": "upload", "error": "Selecciona un archivo Excel (.xlsx)."},
        )

    try:
        rows, detected = parse_excel(file_obj)

        # Asegurar defaults mínimos para que SIEMPRE previsualice e importe
        normalized_rows = []
        for r in rows:
            normalized_rows.append(
                {
                    "ubicacion": _clean_text(r.get("ubicacion")),
                    "sector": _clean_text(r.get("sector")),
                    "piso": _clean_text(r.get("piso")),
                    "tipo_de_lugar": _clean_text(r.get("tipo_de_lugar")) or "Sin especificar",
                    "lugar": _clean_text(r.get("lugar")),
                    "categoria": _clean_text(r.get("categoria")) or "Sin categoría",
                    "objeto": _clean_text(r.get("objeto")),
                    "tipo_objeto": _clean_text(r.get("tipo_objeto")),
                    "cantidad": r.get("cantidad") if r.get("cantidad") is not None else 0,
                    "estado": _clean_text(r.get("estado")),
                    "detalle": _clean_text(r.get("detalle")),
                }
            )

        payload = {"objetos_lugar": normalized_rows}

        return render(
            request,
            "excel/carga_masiva.html",
            {"step": "preview", "payload": payload, "detected": detected},
        )

    except Exception as e:
        return render(
            request,
            "excel/carga_masiva.html",
            {"step": "upload", "error": str(e)},
        )


@login_required

def descargar_plantilla_carga_masiva(request):
    xlsx_bytes = build_excel_plantilla_carga_masiva()
    response = HttpResponse(
        xlsx_bytes,
        content_type ="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename= "PLANTILLA_CARGA_MASIVA.xlsx"'
    return response
