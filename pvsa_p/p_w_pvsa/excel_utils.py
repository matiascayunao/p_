from io import BytesIO
from copy import copy

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.workbook.workbook import Workbook as WorkbookType
from openpyxl.worksheet.datavalidation import DataValidation

from .models import ObjetoLugar


THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FILL_TITLE = PatternFill("solid", fgColor="CFE2F3")
FILL_PISO = PatternFill("solid", fgColor="E7F1FF")

FONT_TITLE = Font(bold=True, size=14)
FONT_PISO = Font(bold=True, size=12)
FONT_HDR = Font(bold=True, size=10)
FONT_CELL = Font(size=10)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

FILL_ESTADO_BUENO = PatternFill("solid", fgColor="C6EFCE")
FILL_ESTADO_PENDIENTE = PatternFill("solid", fgColor="FFF2CC")
FILL_ESTADO_MALO = PatternFill("solid", fgColor="F9CBAD")


# ====== columnas visibles (con separadores ocultos entre medio) ======
# Visible: A,C,E,G,I,K,M (7 columnas)
# Ocultas: B,D,F,H,J,L (separadores para que cada columna colapse INDIVIDUAL)
VISIBLE = {
    "CAT": 1, # A
    "OBJ": 3, # C
    "TIP": 5, # E
    "CAN": 7, # G
    "EST": 9, # I
    "DET": 11, # K
    "FEC": 13, # M
}
MAX_COL = 13 # hasta M


def _safe_sheet_name(name: str) -> str:
    bad = [":", "\\", "/", "?", "*", "[", "]"]
    for ch in bad:
        name = name.replace(ch, " ")
    name = " ".join(name.split()).strip()
    return name[:31] if len(name) > 31 else name


def _unique_sheet_name(wb: WorkbookType, base_name: str) -> str:
    name = _safe_sheet_name(base_name)
    if name not in wb.sheetnames:
        return name
    i = 2
    while True:
        suffix = f"({i})"
        cut = 31 - len(suffix)
        candidate = _safe_sheet_name(name[:cut] + suffix)
        if candidate not in wb.sheetnames:
            return candidate
        i += 1


def _set_col_widths(ws):
    # visibles
    ws.column_dimensions["A"].width = 22 # Categoria
    ws.column_dimensions["C"].width = 26 # Objeto
    ws.column_dimensions["E"].width = 22 # Tipo
    ws.column_dimensions["G"].width = 10 # Cantidad
    ws.column_dimensions["I"].width = 12 # Estado
    ws.column_dimensions["K"].width = 35 # Detalle
    ws.column_dimensions["M"].width = 14 # Fecha

    # separadores ocultos (0 ancho / hidden)
    for col in ["B", "D", "F", "H", "J", "L"]:
        ws.column_dimensions[col].width = 2
        ws.column_dimensions[col].hidden = True


def _style_row(ws, row, c1, c2, fill=None, font=None, align=None):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = BORDER
        if fill:
            cell.fill = copy(fill)
        if font:
            cell.font = copy(font)
        if align:
            cell.alignment = copy(align)


def _setup_outlines(ws):
    # botones de colapsar (filas + columnas)
    ws.sheet_view.showOutlineSymbols = True
    ws.sheet_properties.outlinePr.summaryBelow = False # resumen arriba (PISO/LUGAR)
    ws.sheet_properties.outlinePr.summaryRight = False # símbolo hacia la izquierda (columnas)
    ws.sheet_format.outlineLevelRow = 2
    ws.sheet_format.outlineLevelCol = 1

    # cada columna visible se colapsa INDIVIDUAL (grupos separados por columnas ocultas)
    ws.column_dimensions.group("A", "A", outline_level=1, hidden=False)
    ws.column_dimensions.group("C", "C", outline_level=2, hidden=False)
    ws.column_dimensions.group("E", "E", outline_level=3, hidden=False)
    ws.column_dimensions.group("G", "G", outline_level=4, hidden=False)
    ws.column_dimensions.group("I", "I", outline_level=5, hidden=False)
    ws.column_dimensions.group("K", "K", outline_level=6, hidden=False)
    ws.column_dimensions.group("M", "M", outline_level=7, hidden=False)


def _tipo_txt(ol) -> str:
    marca = (ol.tipo_de_objeto.marca or "").strip()
    material = (ol.tipo_de_objeto.material or "").strip()
    partes = [p for p in (marca, material) if p]
    return " - ".join(partes) if partes else "-"


def _cat_txt(ol) -> str:
    try:
        return (ol.tipo_de_objeto.objeto.objeto_categoria.nombre_de_categoria or "").strip() or "Sin categoría"
    except Exception:
        return "Sin categoría"
    
def _tipo_lugar_lugar_txt(lugar) -> str:
    try:
        return (lugar.lugar_tipo_lugar.tipo_de_lugar or "").strip() or "Sin especificar"
    except Exception:
        return "Sin especificar"


def _set_row_level(ws, row, level: int):
    rd = ws.row_dimensions[row]
    # no bajar niveles si ya tiene (para anidación)
    rd.outline_level = max(int(rd.outline_level or 0), int(level))
    rd.hidden = False


def _write_lugar_block(ws, start_row, lugar, objetos_qs):
    # =========================
    # NUEVO: fila "Tipo de lugar: X" (antes del lugar)
    # =========================
    tipo_row = start_row
    tipo_txt = _tipo_lugar_lugar_txt(lugar)

    ws.merge_cells(start_row=tipo_row, start_column=1, end_row=tipo_row, end_column=MAX_COL)
    t0 = ws.cell(tipo_row, 1, f"Tipo de lugar: {tipo_txt}")
    t0.font = Font(bold=True, size=11)
    t0.alignment = LEFT
    _style_row(ws, tipo_row, 1, MAX_COL, fill=FILL_TITLE, font=Font(bold=True, size=11), align=LEFT)

    # esta fila queda dentro del bloque del piso (nivel 1)
    _set_row_level(ws, tipo_row, 1)

    # =========================
    # LUGAR (fila resumen, colapsa lo de abajo)
    # =========================
    lugar_row = start_row + 1

    ws.merge_cells(start_row=lugar_row, start_column=1, end_row=lugar_row, end_column=MAX_COL)
    t = ws.cell(lugar_row, 1, lugar.nombre_del_lugar)
    t.font = Font(bold=True, size=12)
    t.alignment = CENTER
    _style_row(ws, lugar_row, 1, MAX_COL, fill=FILL_TITLE, font=Font(bold=True, size=12), align=CENTER)

    # nivel 1 para el título de LUGAR (resumen de su grupo)
    _set_row_level(ws, lugar_row, 1)

    # headers (detalle del lugar) -> nivel 2
    hr = lugar_row + 1
    headers = [
        ("Categoría", VISIBLE["CAT"]),
        ("Objeto", VISIBLE["OBJ"]),
        ("Tipo", VISIBLE["TIP"]),
        ("Cantidad", VISIBLE["CAN"]),
        ("Estado", VISIBLE["EST"]),
        ("Detalle", VISIBLE["DET"]),
        ("Fecha", VISIBLE["FEC"]),
    ]
    for label, col in headers:
        cell = ws.cell(hr, col, label)
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.fill = FILL_TITLE
        cell.border = BORDER

    # bordes/estilo en toda la franja (incluye separadores ocultos)
    _style_row(ws, hr, 1, MAX_COL, fill=FILL_TITLE, font=FONT_HDR, align=CENTER)
    _set_row_level(ws, hr, 2)

    r = hr + 1
    count = 0

    for ol in objetos_qs:
        count += 1

        categoria = _cat_txt(ol)
        objeto_nombre = ol.tipo_de_objeto.objeto.nombre_del_objeto
        tipo_txt2 = _tipo_txt(ol)

        ws.cell(r, VISIBLE["CAT"], categoria).font = FONT_CELL
        ws.cell(r, VISIBLE["OBJ"], objeto_nombre).font = FONT_CELL
        ws.cell(r, VISIBLE["TIP"], tipo_txt2).font = FONT_CELL
        ws.cell(r, VISIBLE["CAN"], ol.cantidad).font = FONT_CELL

        estado = ol.get_estado_display()
        estado_cell = ws.cell(r, VISIBLE["EST"], estado)
        estado_cell.font = FONT_CELL

        if estado == "Bueno":
            estado_cell.fill = FILL_ESTADO_BUENO
        elif estado == "Pendiente":
            estado_cell.fill = FILL_ESTADO_PENDIENTE
        elif estado == "Malo":
            estado_cell.fill = FILL_ESTADO_MALO

        ws.cell(r, VISIBLE["DET"], ol.detalle or "-").font = FONT_CELL

        cfecha = ws.cell(r, VISIBLE["FEC"], ol.fecha)
        cfecha.number_format = "dd/mm/yyyy"
        cfecha.font = FONT_CELL

        # bordes en toda la franja (incluye separadores ocultos)
        for c in range(1, MAX_COL + 1):
            ws.cell(r, c).border = BORDER

        # alineación (solo visibles)
        ws.cell(r, VISIBLE["CAT"]).alignment = LEFT
        ws.cell(r, VISIBLE["OBJ"]).alignment = LEFT
        ws.cell(r, VISIBLE["TIP"]).alignment = LEFT
        ws.cell(r, VISIBLE["CAN"]).alignment = CENTER
        ws.cell(r, VISIBLE["EST"]).alignment = CENTER
        ws.cell(r, VISIBLE["DET"]).alignment = LEFT
        ws.cell(r, VISIBLE["FEC"]).alignment = CENTER

        _set_row_level(ws, r, 2)
        r += 1

    if count == 0:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=MAX_COL)
        cell = ws.cell(r, 1, "Sin objetos registrados")
        cell.font = Font(italic=True, size=10)
        cell.alignment = CENTER
        _style_row(ws, r, 1, MAX_COL, align=CENTER)
        _set_row_level(ws, r, 2)
        r += 1

    # fila separadora (parte del piso, no del lugar) -> nivel 1
    _set_row_level(ws, r, 1)
    return r + 1

def build_excel_sectores(ubicaciones_qs):
    wb = Workbook()
    wb.remove(wb.active)

    for ub in ubicaciones_qs:
        sheet_name = _unique_sheet_name(wb, f"{ub.sector.sector} - {ub.ubicacion}")
        ws = wb.create_sheet(title=sheet_name)
        

        _set_col_widths(ws)
        _setup_outlines(ws)

        # Título general
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=MAX_COL)
        ws["A1"] = f"Sector: {ub.sector.sector} | Ubicación: {ub.ubicacion}"
        ws["A1"].font = FONT_TITLE
        ws["A1"].alignment = CENTER
        _style_row(ws, 1, 1, MAX_COL, fill=FILL_TITLE, font=FONT_TITLE, align=CENTER)

        row = 3

        pisos = ub.piso_set.all().order_by("piso")
        for p in pisos:
            piso_row = row

            # PISO (fila resumen)
            ws.merge_cells(start_row=piso_row, start_column=1, end_row=piso_row, end_column=MAX_COL)
            cell = ws.cell(piso_row, 1, f"PISO {p.piso}")
            cell.font = FONT_PISO
            cell.alignment = LEFT
            _style_row(ws, piso_row, 1, MAX_COL, fill=FILL_PISO, font=FONT_PISO, align=LEFT)

            # piso es resumen (nivel 0), sus detalles serán nivel 1/2
            ws.row_dimensions[piso_row].outline_level = 0
            ws.row_dimensions[piso_row].hidden = False

            # fila en blanco dentro del piso (detalle nivel 1)
            row = piso_row + 1
            _set_row_level(ws, row, 1)

            row = piso_row + 2

            lugares = p.lugar_set.all().order_by("id")
            piso_detail_start = piso_row + 1 # todo lo de abajo del piso

            for lugar in lugares:
                objetos = (
                    ObjetoLugar.objects
                    .filter(lugar=lugar)
                    .select_related("tipo_de_objeto__objeto__objeto_categoria")
                    .order_by("id")
                )
                row = _write_lugar_block(ws, row, lugar, objetos)

            piso_detail_end = row - 1
            if piso_detail_end >= piso_detail_start:
                # asegurar nivel 1 mínimo para TODO el bloque del piso
                for rr in range(piso_detail_start, piso_detail_end + 1):
                    _set_row_level(ws, rr, 1)

            row += 1

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def build_excel_plantilla_carga_masiva():
    """
    Plantilla NORMALIZADA para Carga Masiva.
    - Hoja: "ObjetosLugar" (tu parser la busca)
    - Encabezados: ubicacion/sector/lugar/cantidad/estado (y extras)
    - Panel de ayuda en columnas P..V
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "ObjetosLugar"

    # ---- estilos ----
    THIN = Side(style="thin", color="D1D5DB")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    FILL_HDR = PatternFill("solid", fgColor="E5E7EB") # gris claro
    FILL_HELP_T = PatternFill("solid", fgColor="CFE2F3") # celeste
    FILL_HELP_B = PatternFill("solid", fgColor="FFF2CC") # amarillo suave

    FONT_HDR = Font(bold=True, size=10)
    FONT_T = Font(bold=True, size=13)
    FONT_HELP_T = Font(bold=True, size=12)
    FONT_CELL = Font(size=10)

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ---- columnas (A..K) ----
    headers = [
        "Ubicación", # A -> ubicacion
        "Sector", # B -> sector
        "Piso", # C -> piso
        "Tipo de lugar", # D -> tipo_de_lugar
        "Lugar", # E -> lugar
        "Categoría", # F -> categoria
        "Objeto", # G -> objeto
        "Tipo", # H -> tipo_objeto (ej: Marca - Material)
        "Cantidad", # I -> cantidad
        "Estado", # J -> estado (Bueno/Pendiente/Malo)
        "Detalle", # K -> detalle
    ]

    # anchos
    widths = {
        "A": 20, "B": 18, "C": 10, "D": 18, "E": 22,
        "F": 16, "G": 18, "H": 20, "I": 10, "J": 14, "K": 28,
        # panel ayuda
        "P": 18, "Q": 18, "R": 18, "S": 18, "T": 18, "U": 18, "V": 18,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ---- título arriba ----
    ws.merge_cells("A1:K1")
    ws["A1"] = "Plantilla · Carga Masiva (relleno manual)"
    ws["A1"].font = FONT_T
    ws["A1"].alignment = CENTER

    # ---- encabezados (fila 3) ----
    hdr_row = 3
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=hdr_row, column=i, value=h)
        c.font = FONT_HDR
        c.fill = FILL_HDR
        c.alignment = CENTER
        c.border = BORDER

    # borde completo fila header
    for col in range(1, 12):
        ws.cell(hdr_row, col).border = BORDER

    # ---- datos de ejemplo (desde fila 4) ----
    examples = [
        # ubicacion, sector, piso, tipo_lugar, lugar, categoria, objeto, tipo, cantidad, estado, detalle
        ["Edificio Central", "Camino Costero", 1, "Baño", "Baño Hombres", "Sanitario", "Tasas", "Sin marca - Sin material", 2, "Bueno", "OK"],
        ["Edificio Central", "Camino Costero", 1, "Baño", "Baño Hombres", "Higiene", "Dispensadores de papel", "Sin marca - Sin material", 1, "Pendiente", "Falta recarga"],
        ["Edificio Central", "Camino Costero", 1, "Baño", "Baño Hombres", "Infraestructura", "Luces", "Philips - LED", 6, "Bueno", ""],
        ["Carpa", "Camino Costero", 1, "Comedor", "Comedor Principal", "Mobiliario", "Mesas", "Sin marca - Plástico", 10, "Bueno", ""],
        ["Muelle", "Muelle", 1, "Vestidor", "Vestidor 1", "Climatización", "Extractores", "Sin marca - Sin material", 2, "Malo", "No enciende"],
    ]

    start_row = 4
    for r_i, row in enumerate(examples, start=start_row):
        for c_i, val in enumerate(row, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=val)
            cell.font = FONT_CELL
            cell.alignment = LEFT if c_i in (1,2,4,5,6,7,8,11) else CENTER
            cell.border = BORDER

    # ---- validaciones ----
    # Estado: Bueno/Pendiente/Malo
    dv_estado = DataValidation(type="list", formula1='"Bueno,Pendiente,Malo"', allow_blank=True)
    ws.add_data_validation(dv_estado)
    dv_estado.add(f"J{start_row}:J2000")

    # Cantidad: entero >= 0
    dv_cant = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    ws.add_data_validation(dv_cant)
    dv_cant.add(f"I{start_row}:I2000")

    # ---- congelar panel ----
    ws.freeze_panes = "A4"

    # ---- autofiltro ----
    ws.auto_filter.ref = f"A{hdr_row}:K2000"

    # =========================================================
    # Panel de ayuda a la derecha (P..V)
    # =========================================================
    ws.merge_cells("P1:V1")
    ws["P1"] = "Ayuda rápida"
    ws["P1"].font = FONT_HELP_T
    ws["P1"].fill = FILL_HELP_T
    ws["P1"].alignment = CENTER

    help_text = (
        "✅ Una fila = 1 objeto en un lugar.\n\n"
        "OBLIGATORIO:\n"
        "• Ubicación, Sector, Lugar, Cantidad, Estado.\n\n"
        "RECOMENDADO:\n"
        "• Piso: número (ej: 1, 2, 3).\n"
        "• Tipo de lugar: Baño, Comedor, Vestidor, etc.\n"
        "• Categoría / Objeto: según tu catálogo.\n\n"
        "COLUMNA 'Tipo':\n"
        "• Formato sugerido: Marca - Material\n"
        " Ej: Philips - LED\n"
        "• Si no sabes: Sin marca - Sin material\n\n"
        "ESTADO:\n"
        "• Debe ser: Bueno / Pendiente / Malo\n\n"
        "IMPORTANTE:\n"
        "• No cambies los nombres de los encabezados.\n"
        "• Puedes borrar las filas de ejemplo y dejar tus datos.\n"
    )

    ws.merge_cells("P2:V16")
    ws["P2"] = help_text
    ws["P2"].font = FONT_CELL
    ws["P2"].fill = FILL_HELP_B
    ws["P2"].alignment = LEFT

    # bordes panel ayuda
    for r in range(1, 17):
        for col in range(16, 23): # P=16 .. V=22
            ws.cell(r, col).border = BORDER

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


